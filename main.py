from tkinter import *
from tkinter import ttk
import tkinter.messagebox as tmsg
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import date, timedelta
import openpyxl.drawing.image
import time
import json
import os.path
from PIL import Image, ImageTk
from itertools import count
import win32com.client
import os
from os.path import expanduser

# ===================Python Variables=======================

menu_category = ["Laundry", "Steam Ironing", "Dry Cleaning", "Shoe Washing"]

menu_category_dict = {"Laundry": "1 Laundry.txt", "Steam Ironing": "2 Steam Ironing.txt",
                      "Dry Cleaning": "3 Dry Cleaning.txt", "Shoe Washing": "4 Shoe Washing.txt"}
v = os.path.abspath(os.getcwd())

order_dict = {}
for i in menu_category:
    order_dict[i] = {}

os.chdir(os.path.dirname(os.path.abspath(__file__)))


# ====================Backend Functions===========================
def load_menu():
    menuCategory.set("")
    menu_table.delete(*menu_table.get_children())
    menu_file_list = os.listdir("Menu")
    for file in menu_file_list:
        f = open(f"Menu\\" + file, "r")
        category = ""
        while True:
            line = f.readline()
            if line == "":
                menu_table.insert('', END, values=["", "", ""])
                break
            elif line == "\n":
                continue
            elif line[0] == '#':
                category = line[1:-1]
                name = "\t\t" + line[:-1]
                price = ""
            elif line[0] == '*':
                name = line[:-1]
                price = ""
            else:
                name = line[:line.rfind(" ")]
                price = line[line.rfind(" ") + 1:-3]

            menu_table.insert('', END, values=[name, price, category])
            print("menu")


def load_order():
    order_table.delete(*order_table.get_children())
    for category in order_dict.keys():
        if order_dict[category]:
            for lis in order_dict[category].values():
                order_table.insert('', END, values=lis)
    update_total_price()
    print("order")


def add_button_operation():
    name = itemName.get()
    rate = itemRate.get()
    category = itemCategory.get()
    quantity = itemQuantity.get()

    if name in order_dict[category].keys():
        tmsg.showinfo("Error", "Item already exist in your order")
        return
    if not quantity.isdigit():
        tmsg.showinfo("Error", "Please Enter Valid Quantity")
        return
    lis = [name, rate, quantity, str(int(rate) * int(quantity)), category]
    order_dict[category][name] = lis


def load_item_from_menu(event):
    cursor_row = menu_table.focus()
    contents = menu_table.item(cursor_row)
    row = contents["values"]

    itemName.set(row[0])
    itemRate.set(row[1])
    itemCategory.set(row[2])
    itemQuantity.set("1")


def load_item_from_order(event):
    cursor_row = order_table.focus()
    contents = order_table.item(cursor_row)
    row = contents["values"]

    itemName.set(row[1])
    itemRate.set(row[2])
    itemQuantity.set(row[2])
    itemCategory.set(row[4])


def show_button_operation():
    category = menuCategory.get()
    if category not in menu_category:
        tmsg.showinfo("Error", "Please select valid Choice")
    else:
        menu_table.delete(*menu_table.get_children())
        f = open(f"Menu\\" + menu_category_dict[category], "r")
        while True:
            line = f.readline()
            if (line == ""):
                break
            if (line[0] == '#' or line == "\n"):
                continue
            if (line[0] == '*'):
                name = "\t" + line[:-1]
                menu_table.insert('', END, values=[name, "", ""])
            else:
                name = line[:line.rfind(" ")]
                price = line[line.rfind(" ") + 1:-3]
                menu_table.insert('', END, values=[name, price, category])


def clear_button_operation():
    itemName.set("")
    itemRate.set("")
    itemQuantity.set("")
    itemCategory.set("")


def cancel_button_operation():
    names = []
    for i in menu_category:
        names.extend(list(order_dict[i].keys()))
    if len(names) == 0:
        tmsg.showinfo("Error", "Your order list is Empty")
        return
    ans = tmsg.askquestion("Cancel Order", "Are You Sure to Cancel Order?")
    if ans == "yes":
        return
    order_table.delete(*order_table.get_children())
    for i in menu_category:
        order_dict[i] = {}
    clear_button_operation()
    update_total_price()


def update_button_operation():
    name = itemName.get()
    rate = itemRate.get()
    category = itemCategory.get()
    quantity = itemQuantity.get()

    if category == "":
        return
    if name not in order_dict[category].keys():
        tmsg.showinfo("Error", name + "is not in your order list")
        return
    if order_dict[category][name][2] == quantity:
        tmsg.showinfo("Error", "No changes in Quantity")
        return
    order_dict[category][name][2] = quantity
    order_dict[category][name][3] = str(int(rate) * int(quantity))
    load_order()


def remove_button_operation():
    name = itemName.get()
    category = itemCategory.get()

    if category == "":
        return
    if name not in order_dict[category].keys():
        tmsg.showinfo("Error", name + " is not in your order list")
        return
    del order_dict[category][name]
    load_order()


def update_total_price():
    price = 0
    for i in menu_category:
        for j in order_dict[i].keys():
            price += int(order_dict[i][j][3])
    if price == 0:
        totalPrice.set("")
    else:
        totalPrice.set("Rs. " + str(price) + "  /-")


def search_contacts():
    customer_name = customerName.get()
    customer_contact = customerContact.get()
    customer_address = customerAddress.get()
    if not customer_contact.isdigit():
        tmsg.showinfo("Error", "Invalid Customer Contact")
        return
    try:
        fid = open(f"Details\\{customer_contact}.json", 'r')
    except FileNotFoundError:
        tmsg.showinfo("Error", "Contact not Found")
    else:
        f = open(f"Details\\{customer_contact}.json", "r")
        json_object = json.load(f)
        customerName.set(json_object["name"])
        customerContact.set(json_object["phonenumber"])
        customerAddress.set(json_object["Address"])


def add_new_contact():
    customer_name = customerName.get()
    customer_contact = customerContact.get()
    if customer_name == "" or customer_contact == "":
        tmsg.showinfo("Error", "Customer Details Required")
        return
    if not customerContact.get().isdigit():
        tmsg.showinfo("Error", "Invalid Customer Contact")
        return
    dictionary = {
        "name": f"{customer_name}",
        "Address": f"{customerAddress.get()}",
        "phonenumber": f"{customer_contact}"
    }
    add_contact = json.dumps(dictionary, indent=3)
    with open("Details\\" + f"{customer_contact}.json", "w") as outfile:
        outfile.write(add_contact)
        tmsg.showinfo("Success", "Contact Added to List")


class ImageLabel(Label):
    """a label that displays images, and plays them if they are gifs"""

    def load(self, im):
        if isinstance(im, str):
            im = Image.open(im)
        self.loc = 0
        self.frames = []

        try:
            for i in count(1):
                self.frames.append(ImageTk.PhotoImage(im.copy()))
                im.seek(i)
        except EOFError:
            pass

        try:
            self.delay = im.info['duration']
        except:
            self.delay = 100

        if len(self.frames) == 1:
            self.config(image=self.frames[0])
        else:
            self.next_frame()

    def unload(self):
        self.config(image="")
        self.frames = None

    def next_frame(self):
        if self.frames:
            self.loc += 1
            self.loc %= len(self.frames)
            self.config(image=self.frames[self.loc])
            self.after(self.delay, self.next_frame)
            if self.loc == 177:
                print(self.loc)
                self.destroy()


def bill_button_operation():
    customer_name = customerName.get()
    customer_contact = customerContact.get()
    names = []
    for i in menu_category:
        names.extend(list(order_dict[i].keys()))
    if len(names) == 0:
        tmsg.showinfo("Error", "Your order list is Empty")
        return
    if customer_name == "" or customer_contact == "":
        tmsg.showinfo("Error", "Customer Details Required")
        return
    if not customerContact.get().isdigit():
        tmsg.showinfo("Error", "Invalid Customer Contact")
        return
    ans = tmsg.askquestion("Generate Bill", "Are You Sure to Generate Bill?")
    if ans == "yes":
        folder = "Bill Records"
        if not os.path.exists(f"{folder}"):
            os.makedirs(f"{folder}")
        t = time.localtime(time.time())

        with open("Details\\bill_number.json", "r") as outfile:
            json_object = json.load(outfile)
            bill_letter = str(json_object["billLetter"])
            bill_no = int(json_object["billNo"])
        bill_number = str(f"{bill_letter}{bill_no}")
        # Excel File Starting
        xlWorkBook = load_workbook(f"Images\\Workbook_2.xlsx")
        xlSheet = xlWorkBook.active
        del xlWorkBook['Invoice']._images[0]
        xlSheet.title = "Invoice"
        headerxl = ["No.", "Name", "Rate", "Quantity", "Price", "Total Price:- ",  # + totalPrice.get()
                    str(date.today())]
        delivery_date = str(date.today() + timedelta(4))
        col = 0
        row = 0
        img = openpyxl.drawing.image.Image('Images\\Logo.png')
        img.anchor = "F1"
        img.width = 213.165312
        img.height = 91.0866144
        xlSheet.add_image(img)
        xlSheet.cell(row=row + 9, column=col + 1, value=f"Customer Name:- {customerName.get()}")
        xlSheet.cell(row=row + 10, column=col + 1, value=f"Customer Contact No:- {customerContact.get()}")
        xlSheet.cell(row=row + 12, column=col + 1, value=customerAddress.get())
        xlSheet.cell(row=row + 9, column=col + 5, value=headerxl[6])
        xlSheet.cell(row=row + 6, column=col + 7, value=headerxl[6])
        xlSheet.cell(row=row + 6, column=col + 5, value=str(bill_number))
        xlSheet.cell(row=row + 11, column=col + 5, value=str(delivery_date))
        xlSheet.cell(row=row + 32, column=col + 7, value=totalPrice.get())
        xlSheet.print_area = 'A1:G42'
        xlSheet.print_options.verticalCentered = True
        xlSheet.print_options.horizontalCentered = True
        row1 = 17

        for i in menu_category:
            for j in order_dict[i].keys():
                lis = order_dict[i][j]
                # for row_num, data in dict(lis):
                names1 = lis[0]
                rates1 = lis[1]
                quantitys1 = lis[2]
                prices1 = lis[3]
                xlSheet.cell(row=row1, column=col + 1, value=names1)
                xlSheet.cell(row=row1, column=col + 5, value=quantitys1)
                xlSheet.cell(row=row1, column=col + 6, value=rates1)
                xlSheet.cell(row=row1, column=col + 7, value=prices1)
                row1 = row1 + 1

        xlWorkBook.save(f"Bill Records\\{customerName.get()}{bill_number}.xlsx")
        xlWorkBook.close()
        tmsg.showinfo("Success", f"Bill Generated for {bill_number}")

        folder = "Monthly Bills"
        if not os.path.exists(f"{folder}"):
            os.makedirs(f"{folder}")

        if not os.path.exists(f"{folder}"):
            os.makedirs(f"{folder}")
        if not os.path.exists(f"{folder}\\{t.tm_mon},{t.tm_year}.xlsx"):
            print("File not exist")

            with open("Details\\bill_number.json", "r") as outfile:
                json_object = json.load(outfile)
                bill_letter = str(json_object["billLetter"])
                bill_no = int(json_object["billNo"])

            xlMonthlyBills = Workbook()

            xlMonthlyBillsSheet = xlMonthlyBills.active
            xlMonthlyBillsSheet.title = f"{t.tm_mon},{t.tm_year}Bills"
            xlMonthlyBillsSheet.column_dimensions['A'].width = 3
            xlMonthlyBillsSheet.column_dimensions['B'].width = 35
            xlMonthlyBillsSheet.column_dimensions['C'].width = 10
            xlMonthlyBillsSheet.column_dimensions['D'].width = 10
            xlMonthlyBillsSheet.column_dimensions['E'].width = 35
            xlMonthlyBillsSheet.column_dimensions['F'].width = 10
            xlMonthlyBillsSheet.column_dimensions['G'].width = 35
            xlMonthlyBillsSheet.column_dimensions['H'].width = 10
            xlMonthlyBillsSheet.column_dimensions['I'].width = 14
            xlMonthlyBillsSheet.column_dimensions['J'].width = 19
            col1 = 0
            row1 = 0
            number = 1
            xlMonthlyBills_header = ["No.", "Name", "Bill No.", "Mo No.", "Address", "Rec.Date", "Price", "Laundry",
                                     "Dry Clean", "Del. Date", "Paid/Not Paid", "Deliver/Not Deliver"]
            delivery_date = date.today() + timedelta(4)

            a = xlMonthlyBillsSheet.cell(row=row1 + 1, column=col1 + 1, value=xlMonthlyBills_header[0])
            b = xlMonthlyBillsSheet.cell(row=row1 + 1, column=col1 + 2, value=xlMonthlyBills_header[1])
            c = xlMonthlyBillsSheet.cell(row=row1 + 1, column=col1 + 3, value=xlMonthlyBills_header[2])
            d = xlMonthlyBillsSheet.cell(row=row1 + 1, column=col1 + 4, value=xlMonthlyBills_header[3])
            e = xlMonthlyBillsSheet.cell(row=row1 + 1, column=col1 + 5, value=xlMonthlyBills_header[4])
            f = xlMonthlyBillsSheet.cell(row=row1 + 1, column=col1 + 6, value=xlMonthlyBills_header[5])
            g = xlMonthlyBillsSheet.cell(row=row1 + 1, column=col1 + 7, value=xlMonthlyBills_header[6])
            j = xlMonthlyBillsSheet.cell(row=row1 + 1, column=col1 + 8, value=xlMonthlyBills_header[9])
            k = xlMonthlyBillsSheet.cell(row=row1 + 1, column=col1 + 9, value=xlMonthlyBills_header[10])
            l = xlMonthlyBillsSheet.cell(row=row1 + 1, column=col1 + 10, value=xlMonthlyBills_header[11])

            xlMonthlyBillsSheet.cell(row=row1 + 2, column=col1 + 1, value=number)
            xlMonthlyBillsSheet.cell(row=row1 + 2, column=col1 + 2, value=customer_name)
            xlMonthlyBillsSheet.cell(row=row1 + 2, column=col1 + 3, value=bill_number)
            xlMonthlyBillsSheet.cell(row=row1 + 2, column=col1 + 4, value=customer_contact)
            xlMonthlyBillsSheet.cell(row=row1 + 2, column=col1 + 5, value=customerAddress.get())
            xlMonthlyBillsSheet.cell(row=row1 + 2, column=col1 + 6,
                                     value=str(date.today()))
            xlMonthlyBillsSheet.cell(row=row1 + 2, column=col1 + 7, value=totalPrice.get())
            xlMonthlyBillsSheet.cell(row=row1 + 2, column=col1 + 8,
                                     value=str(delivery_date))
            xlMonthlyBillsSheet.cell(row=row1 + 2, column=col1 + 9, value="No")
            xlMonthlyBillsSheet.cell(row=row1 + 2, column=col1 + 10, value="No")

            xlMonthlyBills.save(f"{folder}\\{t.tm_mon},{t.tm_year}.xlsx")
            bill_no += 1
            number += 1
            row1 += 1

            bill = {
                "number": number,
                "row": row1,
                "column": 0,
                "billNo": bill_no,
                "billLetter": "WF"}
            bill_json = json.dumps(bill, indent=5)
            with open("Details\\" + "bill_number.json", "w") as outfile:
                outfile.write(bill_json)
                print("bill number updated.")

        else:
            print("file already exist.")
            with open("Details\\bill_number.json", "r") as outfile:
                json_object = json.load(outfile)
                bill_letter = str(json_object["billLetter"])
                bill_no = int(json_object["billNo"])
                row1 = int(json_object["row"])
                col1 = int(json_object["column"])
                number = int(json_object["number"])

            bill_number = f"{bill_letter}{bill_no}"
            xlMonthlyBills = load_workbook(filename=f"{folder}\\{t.tm_mon},{t.tm_year}.xlsx")
            xlMonthlyBillsSheet = xlMonthlyBills.active
            xlMonthlyBillsSheet.title = f"{t.tm_mon},{t.tm_year}Bills"
            xlMonthlyBillsSheet.column_dimensions['A'].width = 3
            xlMonthlyBillsSheet.column_dimensions['B'].width = 35
            xlMonthlyBillsSheet.column_dimensions['C'].width = 10
            xlMonthlyBillsSheet.column_dimensions['D'].width = 10
            xlMonthlyBillsSheet.column_dimensions['E'].width = 35
            xlMonthlyBillsSheet.column_dimensions['F'].width = 10
            xlMonthlyBillsSheet.column_dimensions['G'].width = 15
            xlMonthlyBillsSheet.column_dimensions['H'].width = 10
            xlMonthlyBillsSheet.column_dimensions['I'].width = 14
            xlMonthlyBillsSheet.column_dimensions['J'].width = 19
            xlMonthlyBills_header = ["No.", "Name", "Bill No.", "Mo No.", "Address", "Rec.Date", "Price", "Del. Date",
                                     "Paid/Not Paid", "Deliver/Not Deliver"]
            delivery_date = date.today() + timedelta(4)

            xlMonthlyBillsSheet.cell(row=row1 + 2, column=col1 + 1, value=number)
            xlMonthlyBillsSheet.cell(row=row1 + 2, column=col1 + 2, value=customer_name)
            xlMonthlyBillsSheet.cell(row=row1 + 2, column=col1 + 3, value=bill_number)
            xlMonthlyBillsSheet.cell(row=row1 + 2, column=col1 + 4, value=customer_contact)
            xlMonthlyBillsSheet.cell(row=row1 + 2, column=col1 + 5, value=customerAddress.get())
            xlMonthlyBillsSheet.cell(row=row1 + 2, column=col1 + 6,
                                     value=str(date.today()))
            xlMonthlyBillsSheet.cell(row=row1 + 2, column=col1 + 7, value=str(totalPrice.get()))
            xlMonthlyBillsSheet.cell(row=row1 + 2, column=col1 + 8,
                                     value=str(delivery_date))
            xlMonthlyBillsSheet.cell(row=row1 + 2, column=col1 + 9, value="No")
            xlMonthlyBillsSheet.cell(row=row1 + 2, column=col1 + 10, value="No")
            xlMonthlyBills.save(f"{folder}\\{t.tm_mon},{t.tm_year}.xlsx")

            bill_no += 1
            number += 1
            row1 += 1

            bill = {
                "number": number,
                "row": row1,
                "column": col1,
                "billNo": bill_no,
                "billLetter": "WF"}
            bill_json = json.dumps(bill, indent=5)
            with open("Details\\" + "bill_number.json", "w") as outfile:
                outfile.write(bill_json)
                print("bill number updated.")

        # Clear operations
        clear_button_operation()
        update_total_price()
        customerName.set("")
        customerContact.set("")
        order_table.delete(*order_table.get_children())
        for i in menu_category:
            order_dict[i] = {}


def payment_button_operation():
    customer_name = customerName.get()
    customer_contact = customerContact.get()
    t = time.localtime(time.time())
    if customer_name == "" or customer_contact == "":
        tmsg.showinfo("Error", "Customer Details Required")
        return
    if not customerContact.get().isdigit():
        tmsg.showinfo("Error", "Invalid Customer Contact")
        return
    ans = tmsg.askquestion("Proceed To Payment", "Are You Sure to Pay Bill?")
    if ans == "yes":

        def find_bill():
            t = time.localtime(time.time())
            xlMonthlyBills = load_workbook(filename=f"Monthly Bills\\{t.tm_mon},{t.tm_year}.xlsx")
            xlMonthlyBillsSheet = xlMonthlyBills.active
            all_rows = list(xlMonthlyBillsSheet.rows)
            bi = billNumber_var.get()
            try:
                xlWorkBook = load_workbook(f"Bill Records\\{customer_name}{bi}.xlsx")
            except FileNotFoundError:
                tmsg.showerror("Error", f"No Bill Found on {customerName.get()}'s Name.")
                return
            xlSheet = xlWorkBook.active
            xlSheet.title = "Invoice"
            del xlWorkBook['Invoice']._images[0]
            xlSheet.cell(row=15, column=5, value="Yes")
            img = openpyxl.drawing.image.Image('Images\\Logo.png')
            img.anchor = "F1"
            img.width = 213.165312
            img.height = 91.0866144
            xlSheet.add_image(img)
            xlSheet.print_area = 'A1:G42'
            xlSheet.print_options.verticalCentered = True
            xlSheet.print_options.horizontalCentered = True
            xlWorkBook.save(f"Bill Records\\{customerName.get()}{bi}.xlsx")
            xlWorkBook.close()
            try:
                wb = load_workbook(f"Monthly Bills\\{t.tm_mon},{t.tm_year}.xlsx")
            except FileNotFoundError:
                tmsg.showerror("Error", "No File of this Month")
                return
            sheet = wb.active
            for row in range(2, sheet.max_row + 1):
                pay_number = sheet['C' + str(row)].value
                if pay_number == bi:
                    sheet.cell(row=row, column=9, value="Yes")
                    wb.save(f"Monthly Bills\\{t.tm_mon},{t.tm_year}.xlsx")
                    tmsg.showinfo("Success", f"Bill Paid for {bi}")
            bill_no_root.destroy()

            # Clear operations
            order_table.delete(*order_table.get_children())
            for i in menu_category:
                order_dict[i] = {}
            clear_button_operation()
            update_total_price()
            customerName.set("")
            customerContact.set("")

        bill_no_root = Toplevel()
        bill_no_root.title("Payment")
        bill_no_root.geometry("670x500+300+100")
        bill_no_root.wm_iconbitmap("Images\\Coffee.ico")
        billNumber = Label(bill_no_root, text="What is the Bill Number?")
        billNumber.pack()
        billNumber_var = StringVar()
        billNumber_var.set("")
        billNumber_entry = Entry(bill_no_root, width=20, font="arial 15", bd=5,
                                 textvariable=billNumber_var)
        billNumber_entry.pack()
        billNumber_button = Button(bill_no_root, text="Find", command=find_bill)
        billNumber_button.pack()
        bill_no_root.focus_set()
        bill_no_root.protocol("WM_DELETE_WINDOW")


def deliver_button_operation():
    customer_name = customerName.get()
    customer_contact = customerContact.get()
    t = time.localtime(time.time())
    if customer_name == "" or customer_contact == "":
        tmsg.showinfo("Error", "Customer Details Required")
        return
    if not customerContact.get().isdigit():
        tmsg.showinfo("Error", "Invalid Customer Contact")
        return
    ans = tmsg.askquestion("Delivery Complete", "Are You Sure to Generate Bill?")
    if ans == "yes":
        def find_bill_deliver():
            t = time.localtime(time.time())
            xlMonthlyBills = load_workbook(filename=f"Monthly Bills\\{t.tm_mon},{t.tm_year}.xlsx")
            xlMonthlyBillsSheet = xlMonthlyBills.active
            all_rows = list(xlMonthlyBillsSheet.rows)
            bi = billNumber_var.get()
            try:
                wb = load_workbook(f"Monthly Bills\\{t.tm_mon},{t.tm_year}.xlsx")
            except FileNotFoundError:
                tmsg.showerror("Error", "No File of this Month")
                return
            sheet = wb.active
            try:
                xlWorkBook = load_workbook(f"Bill Records\\{customer_name}{bi}.xlsx")
            except FileNotFoundError:
                tmsg.showerror("Error", f"No Bill Found on {customerName.get()}'s Name.")
                return
            xlSheet = xlWorkBook.active
            xlSheet.title = "Invoice"
            del xlWorkBook['Invoice']._images[0]
            xlSheet.cell(row=13, column=5, value="Yes")
            xlSheet.cell(row=11, column=5, value=str(date.today()))
            img = openpyxl.drawing.image.Image('Images\\Logo.png')
            img.anchor = "F1"
            img.width = 213.165312
            img.height = 91.0866144
            xlSheet.add_image(img)
            xlSheet.print_area = 'A1:G42'
            xlSheet.print_options.verticalCentered = True
            xlSheet.print_options.horizontalCentered = True
            xlWorkBook.save(f"Bill Records\\{customerName.get()}{bi}.xlsx")
            xlWorkBook.close()
            for row in range(2, sheet.max_row + 1):
                pay_number = sheet['C' + str(row)].value
                if pay_number == bi:
                    sheet.cell(row=row, column=10, value="Yes")
                    sheet.cell(row=row, column=8, value=str(date.today()))
                    wb.save(f"Monthly Bills\\{t.tm_mon},{t.tm_year}.xlsx")
                    if not os.path.exists(f"Monthly Bills\\{t.tm_mon},{t.tm_year}.xlsx"):
                        tmsg.showerror("Error", "No File of this Month")
            tmsg.showinfo("Success", f"Bill Delivered for {bi}")
            # Clear operations
            order_table.delete(*order_table.get_children())
            for i in menu_category:
                order_dict[i] = {}
            # Clear operations
            order_table.delete(*order_table.get_children())
            for i in menu_category:
                order_dict[i] = {}
            clear_button_operation()
            update_total_price()
            customerName.set("")
            customerContact.set("")

        bill_no_root = Toplevel()
        bill_no_root.title("Delivery")
        bill_no_root.geometry("670x500+300+100")
        bill_no_root.wm_iconbitmap("Images\\Coffee.ico")
        billNumber = Label(bill_no_root, text="What is the Bill Number?")
        billNumber.pack()
        billNumber_var = StringVar()
        billNumber_var.set("")
        billNumber_entry = Entry(bill_no_root, width=20, font="arial 15", bd=5,
                                 textvariable=billNumber_var)
        billNumber_entry.pack()
        billNumber_button = Button(bill_no_root, text="Deliver", command=find_bill_deliver)
        billNumber_button.pack()
        bill_no_root.focus_set()
        bill_no_root.protocol("WM_DELETE_WINDOW")


def print_button_operation():
    customer_name = "customerName.get()"
    customer_contact = "9824983757"
    t = time.localtime(time.time())
    if customer_name == "" or customer_contact == "":
        tmsg.showinfo("Error", "Customer Details Required")
        return
    if not customer_contact.isdigit():
        tmsg.showinfo("Error", "Invalid Customer Contact")
        return

    def print_bill():
        t = time.localtime(time.time())
        v = os.path.abspath(os.getcwd())
        o = win32com.client.Dispatch('Excel.Application')
        """o.Visible = True
        o.Interactive = True"""
        wb = o.Workbooks.Open(f'{v}/Bill Records/{customerName.get()}{billNumber_var.get()}.xlsx')
        ws = wb.Worksheets("Invoice")
        ws.PrintOut()
        wb.Close()
        o.Application.Quit()
        del o, wb, ws
        tmsg.showinfo("Success", f"Bill Printed for {billNumber_var.get()}")
        bill_no_root.destroy()
        # Clear operations
        order_table.delete(*order_table.get_children())
        for i in menu_category:
            order_dict[i] = {}
        clear_button_operation()
        update_total_price()
        customerName.set("")
        customerContact.set("")

    bill_no_root = Toplevel()
    bill_no_root.title("Print")
    bill_no_root.geometry("670x500+300+100")
    bill_no_root.wm_iconbitmap("Images\\Coffee.ico")
    billNumber = Label(bill_no_root, text="What is the Bill Number?")
    billNumber.pack()
    billNumber_var = StringVar()
    billNumber_var.set("")
    billNumber_entry = Entry(bill_no_root, width=20, font="arial 15", bd=5,
                             textvariable=billNumber_var)

    billNumber_entry.pack()
    billNumber_button = Button(bill_no_root, text="Print", command=print_bill)
    billNumber_button.pack()


def close_window(bill):
    tmsg.showinfo("Thanks", "Thanks for using our service")
    bill.protocol("WM_DELETE_WINDOW")


# ==================Backend Code Ends===============

# ================Frontend Code Start==============

"""
Loading = Tk()
lbl = ImageLabel(Loading)
lbl.load('Images\\SpyderWeb Technologies.gif')
lbl.pack()
Loading.after(6000, Loading.destroy)
Loading.mainloop()
"""
root = Tk()
w, h = root.winfo_screenwidth(), root.winfo_screenheight()
root.geometry("%dx%d+0+0" % (w, h))
root.title("Welcome to Techware Billing System")
root.wm_iconbitmap("Images\\Burger.ico")
# root.attributes('-fullscreen', True)
# root.resizable(0, 0)

# ================Title==============
style_button = ttk.Style()
style_button.configure("TButton", font=("arial", 10, "bold"),
                       background="lightgreen")

title_frame = Frame(root, bd=8, bg="yellow", relief=GROOVE)
title_frame.pack(side=TOP, fill="x")

title_label = Label(title_frame, text="TechWare Billing System",
                    font=("times new roman", 20, "bold"), bg="yellow", fg="red", pady=5)
title_label.pack()

# ==============Customer=============
customer_frame = LabelFrame(root, text="Customer Details", font=("times new roman", 15, "bold"),
                            bd=8, bg="lightblue", relief=GROOVE)
customer_frame.pack(side=TOP, fill="x")

search_button_image = PhotoImage(file="Images\\Search Icon.png")

contact_button_image = PhotoImage(file="Images\\Add Button.png")

customer_name_label = Label(customer_frame, text="Name",
                            font=("arial", 15, "bold"), bg="lightblue", fg="blue")
customer_name_label.grid(row=0, column=0)

customerName = StringVar()
customerName.set("")
customer_name_entry = Entry(customer_frame, width=20, font="arial 15", bd=5,
                            textvariable=customerName)
customer_name_entry.grid(row=0, column=1, padx=50)

customer_contact_label = Label(customer_frame, text="Contact",
                               font=("arial", 15, "bold"), bg="lightblue", fg="blue")
customer_contact_label.grid(row=0, column=2)

customerContact = StringVar()
customerContact.set("")
customer_contact_entry = Entry(customer_frame, width=20, font="arial 15", bd=5,
                               textvariable=customerContact)
customer_contact_entry.grid(row=0, column=3, padx=25)

search_button = Button(customer_frame, image=search_button_image, width=0,
                       command=search_contacts)
search_button.grid(row=0, column=4)

customer_address_label = Label(customer_frame, text="Address",
                               font=("arial", 15, "bold"), bg="lightblue", fg="blue")
customer_address_label.grid(row=0, column=5, padx=30)

customerAddress = StringVar()
customerAddress.set("")
customer_Address_entry = Entry(customer_frame, width=20, font="arial 15", bd=5,
                               textvariable=customerAddress)
customer_Address_entry.grid(row=0, column=6, padx=25)

contact_button = Button(customer_frame, text="Add Contact", image=contact_button_image, compound=LEFT, width=100,
                        command=add_new_contact)
contact_button.grid(row=0, column=7)

# ===============Menu===============
menu_frame = Frame(root, bd=8, bg="lightgreen", relief=GROOVE)
menu_frame.place(x=0, y=125, height=585, width=680)

menu_label = Label(menu_frame, text="Menu",
                   font=("times new roman", 20, "bold"), bg="lightgreen", fg="red", pady=0)
menu_label.pack(side=TOP, fill="x")

menu_category_frame = Frame(menu_frame, bg="lightgreen", pady=10)
menu_category_frame.pack(fill="x")

combo_lable = Label(menu_category_frame, text="Select Type",
                    font=("arial", 12, "bold"), bg="lightgreen", fg="blue")
combo_lable.grid(row=0, column=0, padx=10)

menuCategory = StringVar()
combo_menu = ttk.Combobox(menu_category_frame, values=menu_category,
                          textvariable=menuCategory)
combo_menu.grid(row=0, column=1, padx=30)

show_button = ttk.Button(menu_category_frame, text="Show", width=10,
                         command=show_button_operation)
show_button.grid(row=0, column=2, padx=60)

show_all_button = ttk.Button(menu_category_frame, text="Show All",
                             width=10, command=load_menu)
show_all_button.grid(row=0, column=3)

############################# Menu Tabel ##########################################
menu_table_frame = Frame(menu_frame)
menu_table_frame.pack(fill=BOTH, expand=1)

scrollbar_menu_x = Scrollbar(menu_table_frame, orient=HORIZONTAL)
scrollbar_menu_y = Scrollbar(menu_table_frame, orient=VERTICAL)

style = ttk.Style()
style.configure("Treeview.Heading", font=("arial", 13, "bold"))
style.configure("Treeview", font=("arial", 12), rowheight=25)

menu_table = ttk.Treeview(menu_table_frame, style="Treeview",
                          columns=("name", "price", "category"), xscrollcommand=scrollbar_menu_x.set,
                          yscrollcommand=scrollbar_menu_y.set)

menu_table.heading("name", text="Name")
menu_table.heading("price", text="Price")
menu_table["displaycolumns"] = ("name", "price")
menu_table["show"] = "headings"
menu_table.column("price", width=50, anchor='center')

scrollbar_menu_x.pack(side=BOTTOM, fill=X)
scrollbar_menu_y.pack(side=RIGHT, fill=Y)

scrollbar_menu_x.configure(command=menu_table.xview)
scrollbar_menu_y.configure(command=menu_table.yview)

menu_table.pack(fill=BOTH, expand=1)

load_menu()
menu_table.bind("<ButtonRelease-1>", load_item_from_menu)

###########################################################################################

# ===============Item Frame=============
item_frame = Frame(root, bd=8, bg="lightgreen", relief=GROOVE)
item_frame.place(x=680, y=125, height=230, width=680)

item_title_label = Label(item_frame, text="Item",
                         font=("times new roman", 20, "bold"), bg="lightgreen", fg="red")
item_title_label.pack(side=TOP, fill="x")

item_frame2 = Frame(item_frame, bg="lightgreen")
item_frame2.pack(fill=X)

itemCategory = StringVar()
itemCategory.set("")

itemName = StringVar()
itemName.set("")
item_name = Entry(item_frame2, font="arial 12", textvariable=itemName, state=DISABLED, width=25)
item_name.grid(row=0, column=1, padx=10)

item_rate_label = Label(item_frame2, text="Rate",
                        font=("arial", 12, "bold"), bg="lightgreen", fg="blue")
item_rate_label.grid(row=0, column=2, padx=40)

itemRate = StringVar()
itemRate.set("")
item_rate = Entry(item_frame2, font="arial 12", textvariable=itemRate, width=10)
item_rate.grid(row=0, column=3, padx=10)

item_quantity_label = Label(item_frame2, text="Quantity",
                            font=("arial", 12, "bold"), bg="lightgreen", fg="blue")
item_quantity_label.grid(row=1, column=0, padx=30, pady=15)

itemQuantity = StringVar()
itemQuantity.set("")
item_quantity = Entry(item_frame2, font="arial 12", textvariable=itemQuantity, width=10)
item_quantity.grid(row=1, column=1)

item_frame3 = Frame(item_frame, bg="lightgreen")
item_frame3.pack(fill=X)

add_button = ttk.Button(item_frame3, text="Add Item"
                        , command=add_button_operation)
add_button.grid(row=0, column=0, padx=40, pady=30)

remove_button = ttk.Button(item_frame3, text="Remove Item"
                           , command=remove_button_operation)
remove_button.grid(row=0, column=1, padx=40, pady=30)

update_button = ttk.Button(item_frame3, text="Update Quantity"
                           , command=update_button_operation)
update_button.grid(row=0, column=2, padx=40, pady=30)

clear_button = ttk.Button(item_frame3, text="Clear",
                          width=8, command=clear_button_operation)
clear_button.grid(row=0, column=3, padx=40, pady=30)

# ==============Order Frame=====================
order_frame = Frame(root, bd=8, bg="lightgreen", relief=GROOVE)
order_frame.place(x=680, y=335, height=370, width=680)

order_title_label = Label(order_frame, text="Your Order",
                          font=("times new roman", 20, "bold"), bg="lightgreen", fg="red")
order_title_label.pack(side=TOP, fill="x")

############################## Order Table ###################################
order_table_frame = Frame(order_frame)
order_table_frame.place(x=0, y=40, height=260, width=680)

scrollbar_order_x = Scrollbar(order_table_frame, orient=HORIZONTAL)
scrollbar_order_y = Scrollbar(order_table_frame, orient=VERTICAL)

order_table = ttk.Treeview(order_table_frame,
                           columns=("name", "rate", "quantity", "price", "category"),
                           xscrollcommand=scrollbar_order_x.set,
                           yscrollcommand=scrollbar_order_y.set)

order_table.heading("name", text="Name")
order_table.heading("rate", text="Rate")
order_table.heading('quantity', text="Quantity")
order_table.heading("price", text="Price")
order_table["displaycolumns"] = ("name", "rate", "quantity", "price")
order_table["show"] = "headings"
order_table.column("rate", width=100, anchor='center', stretch=NO)
order_table.column("quantity", width=100, anchor='center', stretch=NO)
order_table.column("price", width=100, anchor='center', stretch=NO)

# order_table.bind("<ButtonRelease-1>", load_item_from_order)

scrollbar_order_x.pack(side=BOTTOM, fill=X)
scrollbar_order_y.pack(side=RIGHT, fill=Y)

scrollbar_order_x.configure(command=order_table.xview)
scrollbar_order_y.configure(command=order_table.yview)

order_table.pack(fill=BOTH, expand=1)

###########################################################################################

total_price_label = Label(order_frame, text="Total Price",
                          font=("arial", 12, "bold"), bg="lightgreen", fg="blue")
total_price_label.pack(side=LEFT, anchor=SW, padx=20, pady=10)

totalPrice = StringVar()
totalPrice.set("")
total_price_entry = Entry(order_frame, font="arial 12", textvariable=totalPrice, state=DISABLED,
                          width=10)
total_price_entry.pack(side=LEFT, anchor=SW, padx=0, pady=10)

bill_button = ttk.Button(order_frame, text="Bill", width=10,
                         command=bill_button_operation)
bill_button.pack(side=LEFT, anchor=SW, padx=5, pady=10)

payment_button = ttk.Button(order_frame, text="Payment", width=10,
                            command=payment_button_operation)
payment_button.pack(side=LEFT, anchor=SW, padx=5, pady=10)

deliver_button = ttk.Button(order_frame, text="Delivered", width=10,
                            command=deliver_button_operation)
deliver_button.pack(side=LEFT, anchor=SW, padx=5, pady=10)

print_button = ttk.Button(order_frame, text="Print", command=print_button_operation, width=10)

print_button.pack(side=LEFT, anchor=SW, padx=5, pady=10)

cancel_button = ttk.Button(order_frame, text="Cancel", command=cancel_button_operation, width=10)

cancel_button.pack(side=LEFT, anchor=SW, padx=5, pady=10)

root.mainloop()

# ====================Frontend code ends===================== #


# ====================Login Page============================= #
