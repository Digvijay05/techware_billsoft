# Getting Details From Database
import json
import re
import sqlite3
from datetime import date, timedelta

import openpyxl.drawing.image
from openpyxl import load_workbook

# Connecting To Database
conn = sqlite3.connect("DB\\Business.db")
cursor = conn.cursor()

# Executing Commands In Cursor
cursor.execute("SELECT * FROM CATEGORY")

# Rows of Data
rows = cursor.fetchall()

# Initializing Items List
items = {}

# Through For Loop
for id, category in rows:
    # Appending To Items List
    items[category] = {}

# Initializing Categories List
categories = {}

# Through For Loop
for id, category in rows:
    # Appending To Categories List
    categories[category] = {}

conn = sqlite3.connect("DB\\Invoices.db")
cursor = conn.cursor()
cursor.execute("SELECT * FROM INVOICES WHERE Invoice_No = ?", ("WF1001",))
rows = cursor.fetchone()
total = re.findall(r"[^\W\d_]+|\d+", rows[14])
total_amount = total[1]
payment = re.findall(r"[^\W\d_]+|\d+", rows[18])
payment_d = payment[0]
payment_due = int(total_amount) - int(payment_d)
# Excel File Starting
xlWorkBook = load_workbook(f"Images\\Workbook_2.xlsx")
xlSheet = xlWorkBook.active
xlSheet.title = "Invoice"
del xlWorkBook['Invoice']._images[0]
headerxl = ["No.", "Name", "Rate", "Quantity", "Price", "Total Price:- ",  # + totalPrice.get()
            str(date.today())]
delivery_date = str(date.today() + timedelta(4))
col = 0
row = 0
img = openpyxl.drawing.image.Image('Images\\Logo.png')
img.anchor = "F1"
img.width = 213.165312
img.height = 91.0866144
xlSheet.add_image(img)
xlSheet.cell(row=row + 9, column=col + 1, value=f"Customer Name:- {rows[7]}")
xlSheet.cell(row=row + 10, column=col + 1, value=f"Customer Contact No:- {rows[6]}")
xlSheet.cell(row=row + 12, column=col + 1, value=rows[8])
xlSheet.cell(row=row + 9, column=col + 5, value=headerxl[6])
xlSheet.cell(row=row + 6, column=col + 7, value=headerxl[6])
xlSheet.cell(row=row + 6, column=col + 5, value=rows[2])
xlSheet.cell(row=row + 11, column=col + 5, value=str(delivery_date))
xlSheet.cell(row=row + 13, column=col + 5, value=str("Rs. " + str(payment_due) + "/-"))
xlSheet.cell(row=row + 33, column=col + 7, value=rows[11])
xlSheet.cell(row=row + 34, column=col + 7, value=rows[12])
xlSheet.cell(row=row + 32, column=col + 7, value=rows[13])
xlSheet.cell(row=row + 35, column=col + 7, value=rows[14])
xlSheet.print_area = 'A1:G43'
xlSheet.print_options.verticalCentered = True
xlSheet.print_options.horizontalCentered = True
row1 = 17
with open(f"Details\\{rows[7]}{rows[2]}.json", "r") as outfile:
    items = json.load(outfile)
print(items)

for i in categories:
    for j in items[i].keys():
        lis = items[i][j]
        names1 = lis[0]
        category = lis[6]
        sub_category = lis[5]
        rates1 = lis[1]
        units1 = lis[11]
        quantitys1 = lis[2]
        prices1 = lis[3]
        xlSheet.cell(row=row1, column=col + 1, value=category + " " + sub_category + " " + names1)
        xlSheet.cell(row=row1, column=col + 5, value=quantitys1)
        xlSheet.cell(row=row1, column=col + 6, value=rates1)
        xlSheet.cell(row=row1, column=col + 4, value=units1)
        xlSheet.cell(row=row1, column=col + 7, value=prices1)
        row1 = row1 + 1
xlWorkBook.save(f"Bill Records\\{rows[7]}{rows[2]}.xlsx")
xlWorkBook.close()
