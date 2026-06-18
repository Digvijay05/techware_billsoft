"""
Invoice file-generation service.

Handles Excel (.xlsx), PDF, and PNG export and win32com printing.
All heavy I/O lives here so that it can be called from a QThread.
"""
from __future__ import annotations

import json
import os
import re
from datetime import date, timedelta
from typing import Dict

import openpyxl.drawing.image
from openpyxl import load_workbook

from src.services import db_service

# Resolve directories relative to the project root
_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
IMAGES_DIR = os.path.join(_PROJECT_ROOT, "Images")
DETAILS_DIR = os.path.join(_PROJECT_ROOT, "Details")
BILL_RECORDS_DIR = os.path.join(_PROJECT_ROOT, "Bill Records")


def _ensure_dirs() -> None:
    os.makedirs(DETAILS_DIR, exist_ok=True)
    os.makedirs(BILL_RECORDS_DIR, exist_ok=True)


# ------------------------------------------------------------------
# JSON item detail persistence
# ------------------------------------------------------------------

def save_items_json(client_name: str, invoice_no: str, items: dict) -> str:
    """Persist the items dict to a JSON file.  Returns the file path."""
    _ensure_dirs()
    path = os.path.join(DETAILS_DIR, f"{client_name}{invoice_no}.json")
    with open(path, "w") as f:
        json.dump(items, f, indent=4)
    return path


def load_items_json(client_name: str, invoice_no: str) -> dict:
    """Load items dict from a previously saved JSON file."""
    path = os.path.join(DETAILS_DIR, f"{client_name}{invoice_no}.json")
    with open(path, "r") as f:
        return json.load(f)


# ------------------------------------------------------------------
# Excel generation (mirrors legacy save_excel)
# ------------------------------------------------------------------

def generate_excel(invoice_no: str, items: dict, categories: list) -> str:
    """
    Create an Excel invoice from the template workbook.

    Parameters
    ----------
    invoice_no : str
        Invoice number used to look up the saved invoice row.
    items : dict
        Nested dict of category → item_key → legacy 12-element list.
    categories : list[str]
        Ordered list of category names.

    Returns
    -------
    str – path to the saved .xlsx file.
    """
    _ensure_dirs()

    # Fetch saved invoice row from DB
    inv = db_service.get_invoice_by_no(invoice_no)
    if not inv:
        raise ValueError(f"Invoice {invoice_no} not found in database.")

    total_parts = re.findall(r"[^\W\d_]+|\d+", str(inv["Total"]))
    total_amount = int(total_parts[1]) if len(total_parts) > 1 else 0
    pay_parts = re.findall(r"[^\W\d_]+|\d+", str(inv["Payment_Amount"]))
    payment_d = int(pay_parts[0]) if pay_parts else 0
    payment_due = total_amount - payment_d

    # Staff info
    staff = db_service.get_staff_by_code(inv["Sold_By"])

    # Load template
    wb = load_workbook(os.path.join(IMAGES_DIR, "Workbook_2.xlsx"))
    ws = wb.active
    ws.title = "Invoice"

    # Remove old embedded image if present
    if ws._images:
        del ws._images[0]

    today_str = str(date.today())
    delivery_date = str(date.today() + timedelta(4))

    # Add logo
    logo = openpyxl.drawing.image.Image(os.path.join(IMAGES_DIR, "Logo.png"))
    logo.anchor = "F1"
    logo.width = 213.165312
    logo.height = 91.0866144
    ws.add_image(logo)

    # Header cells
    ws.cell(row=9,  column=1, value=f"Customer Name:- {inv['Client_Name']}")
    ws.cell(row=10, column=1, value=f"Customer Contact No:- {inv['Client_Contact']}")
    ws.cell(row=12, column=1, value=inv["Client_Address"])
    ws.cell(row=9,  column=5, value=today_str)
    ws.cell(row=6,  column=7, value=today_str)
    ws.cell(row=6,  column=5, value=inv["Invoice_No"])
    ws.cell(row=11, column=5, value=delivery_date)
    ws.cell(row=13, column=5, value=f"Rs. {payment_due}/-")
    ws.cell(row=33, column=7, value=inv["Discount"])
    ws.cell(row=34, column=7, value=inv["Shipping"])
    ws.cell(row=32, column=7, value=inv["SubTotal"])
    ws.cell(row=35, column=7, value=inv["Total"])
    ws.print_area = "A1:G43"
    ws.print_options.verticalCentered = True
    ws.print_options.horizontalCentered = True

    # Item rows (start at Excel row 17)
    row_idx = 17
    for cat in categories:
        if cat not in items:
            continue
        for key, lis in items[cat].items():
            name = lis[0]
            rate = lis[1]
            qty = lis[2]
            price = lis[3]
            sub_cat = lis[5]
            category = lis[6]
            unit = lis[11]
            ws.cell(row=row_idx, column=1, value=f"{category} {sub_cat} {name}")
            ws.cell(row=row_idx, column=4, value=unit)
            ws.cell(row=row_idx, column=5, value=qty)
            ws.cell(row=row_idx, column=6, value=rate)
            ws.cell(row=row_idx, column=7, value=price)
            row_idx += 1

    out_path = os.path.join(BILL_RECORDS_DIR, f"{inv['Client_Name']}{inv['Invoice_No']}.xlsx")
    wb.save(out_path)
    wb.close()
    return out_path


# ------------------------------------------------------------------
# PDF / PNG export via win32com  (Windows-only)
# ------------------------------------------------------------------

def export_pdf(invoice_no: str) -> str:
    """Convert the Excel invoice to PDF using Excel COM automation."""
    import win32com.client
    from win32com.universal import com_error

    inv = db_service.get_invoice_by_no(invoice_no)
    if not inv:
        raise ValueError(f"Invoice {invoice_no} not found.")

    wb_path = os.path.join(BILL_RECORDS_DIR, f"{inv['Client_Name']}{inv['Invoice_No']}.xlsx")
    pdf_path = wb_path.replace(".xlsx", ".pdf")

    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    try:
        wb = excel.Workbooks.Open(os.path.abspath(wb_path))
        wb.WorkSheets(1).Select()
        wb.ActiveSheet.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
    except com_error as e:
        raise RuntimeError(f"PDF export failed: {e}")
    finally:
        try:
            wb.Close(False)
        except Exception:
            pass
        excel.Quit()

    return pdf_path


def print_invoice(invoice_no: str) -> None:
    """Print the Excel invoice via win32com."""
    import win32com.client
    from win32com.universal import com_error

    inv = db_service.get_invoice_by_no(invoice_no)
    if not inv:
        raise ValueError(f"Invoice {invoice_no} not found.")

    wb_path = os.path.join(BILL_RECORDS_DIR, f"{inv['Client_Name']}{inv['Invoice_No']}.xlsx")

    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    try:
        wb = excel.Workbooks.Open(os.path.abspath(wb_path))
        wb.WorkSheets(1).Select()
        wb.ActiveSheet.PrintOut()
        wb.Save()
        wb.Close()
    except com_error as e:
        raise RuntimeError(f"Printing failed: {e}")
    finally:
        excel.Quit()
