# Importing Packages
import json
import os
import sqlite3
import time
import openpyxl.drawing.image
import wand
import wand.image
import win32com
import win32com.client
import re

from datetime import date, timedelta
from tkinter import *
from tkinter import messagebox as tmsg
from tkinter import ttk, scrolledtext
from Digvijay_Algos.custom_widgets import Custom_treeview, Link_Text, Required_Text
from openpyxl import load_workbook
from tkcalendar import DateEntry
from ttkwidgets.autocomplete import *
from wand import exceptions


class Invoice:
    shippingprice: int

    # Initializing The Invoice
    def __init__(self, root):
        # Root Window
        self.invoice_root = root
        self.invoice_root.attributes('-toolwindow', 1)
        # self.invoice_root.attributes('-topmost', 1)

        self.invoice_root.title("Unsaved Invoice")
        root.geometry("%dx%d+66+16" % (1185, 665))
        # self.invoice_root['bg'] = "white"

        # Style Of TTK
        self.style = ttk.Style(self.invoice_root)

        # Loading TTK Themes
        self.invoice_root.tk.eval("""
                set base_theme_dir C:/Users/Digvijay/Downloads/awthemes-10.2.0/awthemes-10.2.0

                package ifneeded awthemes 10.2.0 \
                    [list source [file join $base_theme_dir awthemes.tcl]]
                package ifneeded colorutils 4.8 \
                    [list source [file join $base_theme_dir colorutils.tcl]]
                package ifneeded awdark 7.7 \
                    [list source [file join $base_theme_dir awdark.tcl]]
                package ifneeded awlight 7.9 \
                    [list source [file join $base_theme_dir awlight.tcl]]
                package ifneeded awbreeze 7.9 \
                    [list source [file join $base_theme_dir awbreeze.tcl]]
                """)

        # Load The Awdark And Awlight Themes
        self.invoice_root.tk.call("package", "require", 'awthemes')
        self.invoice_root.tk.call("package", "require", 'awlight')
        # self.invoice_root.tk.call("package", "require", 'awbreeze')

        # Using Theme AWLIGHT
        self.style.theme_use('awlight')

        # Showing All The Themes
        # print(self.style.theme_names())

        # Button Style 1
        self.style.configure("C.TButton", font=("Calibri", 12), takefocus=False,
                             shiftrelief=0)
        self.style.map("C.TButton",
                       foreground=[('!active', 'white'), ('pressed', 'white'), ('active', 'white')],
                       background=[('!active', '#11481a'), ('pressed', '#39d44b'), ('active', '#096119')]
                       )

        # Button Style 2
        self.style.map("S.TButton",
                       foreground=[('!active', 'white'), ('pressed', 'white'), ('active', 'white')],
                       background=[('!active', '#00bd61'), ('pressed', '#00e375'), ('active', '#00d16c')]
                       )

        # Radiobutton style configure
        self.style.configure("S.TRadiobutton", font=("Calibri", 12), bg="white")

        # Radiobutton style map
        self.style.map("S.TRadiobutton",
                       background=[("!active", "SystemButtonFace"), ("active", "lightgray"), ("pressed", "white"), ],
                       foreground=[("!active", "black"), ("active", "black"), ("pressed", "black"), ],
                       highlight_background=[("!active", "yellow"), ("active", "green"), ("pressed", "blue"), ])

        # CheckButton Style
        self.style.configure('Red.TCheckbutton', font=("Calibri", 12), takefocus=False, background="SystemButtonFace",
                             selectcolor="white", lightcolor="white", darkcolor="white", padding=[5, 0, 1, 2],
                             focusthickness=0, borderwidth=0, relief='none')
        self.style.map('Red.TCheckbutton',
                       background=[('!active', 'SystemButtonFace'), ('pressed', 'white'), ('active', 'white')],
                       foreground=[('!active', 'black'), ('pressed', 'black'), ('active', 'black')])

        # Frame Style
        self.style.map('Retail.TFrame',
                       background=[('!active', 'black'), ('pressed', 'white'), ('active', 'white')])

        # Label Style
        self.style.map('Retail.TLabel',
                       background=[('!active', 'black'), ('pressed', 'white'), ('active', 'white')])

        # Invoice Information Frame
        self.invoice_information_lbl = LabelFrame(self.invoice_root, text="Invoice Information", )  # bg="white")
        self.invoice_information_lbl.place(relx=0.005, rely=0, relwidth=0.99, relheight=0.20)

        # Variable for Contact No.
        self.contact_no_var = StringVar()

        # Variable for Contact Address
        self.contact_address_var = StringVar()

        # Invoice Type Label
        self.invoice_type_lbl = Required_Text(self.invoice_information_lbl, required_text="Invoice Type   ")
        # bg="white")
        self.invoice_type_lbl.required_frame.place(relx=0.025, rely=0)

        # Invoice Type ComboBox
        self.invoice_type_txt = ttk.Combobox(self.invoice_information_lbl, font=("Calibri", 12),
                                             values=["Select", "GST",
                                                     "Non-GST", "Bill of Supply"])
        self.invoice_type_txt.place(relx=0.025, rely=0.2, relwidth=0.17)
        self.invoice_type_txt.current(0)

        # Invoice No. Variable
        self.invoice_no_var = StringVar()

        # Invoice No. Label
        self.invoice_no_lbl = Label(self.invoice_information_lbl, text="Invoice-No.",
                                    font=("Calibri", 10), )  # bg="white")
        self.invoice_no_lbl.place(relx=0.22, rely=0)

        # Invoice No. Entry
        self.invoice_no_txt = ttk.Entry(self.invoice_information_lbl, font=("Calibri", 12),
                                        textvariable=self.invoice_no_var)
        self.invoice_no_txt.place(relx=0.22, rely=0.2, relwidth=0.17)

        # Adding Invoice No. Dynamically
        self.conn = sqlite3.connect("DB\\Business.db")
        self.cursor = self.conn.cursor()
        self.sql_string = '''SELECT * FROM ID where Category = ?'''
        self.cursor.execute(self.sql_string, ("Invoice",))
        self.rows = self.cursor.fetchall()

        self.invoice_no = self.rows[0][1]
        self.invoice_no_var.set(self.invoice_no)

        # Date Label
        self.invoice_date_lbl = Label(self.invoice_information_lbl, text="Date", font=("Calibri", 10), )  # bg="white")
        self.invoice_date_lbl.place(relx=0.415, rely=0)

        # Date Entry
        self.invoice_date_txt = DateEntry(self.invoice_information_lbl, font=("Calibri", 12))
        self.invoice_date_txt.place(relx=0.415, rely=0.2, relwidth=0.17)

        # Place Of Supply Label
        self.invoice_place_of_supply_lbl = Label(self.invoice_information_lbl, text="Place of Supply",
                                                 font=("Calibri", 10), )  # bg="white")
        self.invoice_place_of_supply_lbl.place(relx=0.61, rely=0)

        # Place Of Supply Combobox
        self.invoice_place_of_supply_txt = AutocompleteCombobox(
            self.invoice_information_lbl,
            completevalues=["Select",
                            "Andaman and Nicobar",
                            "Andhra Pradesh",
                            "Arunachal Pradesh",
                            "Assam",
                            "Bihar",
                            "Chandigarh",
                            "Chhattisgarh",
                            "Dadra and Nagar Haveli",
                            "Daman and Diu",
                            "Delhi",
                            "Goa",
                            "Gujarat",
                            "Haryana",
                            "Himachal Pradesh",
                            "Jammu & Kashmir",
                            "Jharkhand",
                            "Karnataka",
                            "Kerala",
                            "Lakshadweep",
                            "Madhya Pradesh",
                            "Maharashtra",
                            "Manipur",
                            "Meghalaya",
                            "Mizoram",
                            "Nagaland",
                            "Odisha",
                            "Puducherry",
                            "Punjab",
                            "Rajashthan",
                            "Sikkim",
                            "Tamil Nadu",
                            "Telangana",
                            "Tripura",
                            "Uttar Pradesh",
                            "Uttarakhand",
                            "West Bengal"], font=("Calibri", 12))
        self.invoice_place_of_supply_txt.current(29)
        self.invoice_place_of_supply_txt.place(relx=0.61, rely=0.2, relwidth=0.17)

        # Bill To Label
        self.bill_to_lbl = Label(self.invoice_information_lbl, text="Bill To", font=("Calibri", 10), )  # bg="white")
        self.bill_to_lbl.place(relx=0.805, rely=0)

        # Client A/c / Cash A/c
        self.bill_to_var = StringVar()
        self.bill_to_var.set("cash")

        # Client A/c RadioButton
        self.client_ac_radio = ttk.Radiobutton(self.invoice_information_lbl, text="Client A/c", value="client",
                                               style="S.TRadiobutton",
                                               variable=self.bill_to_var, command=self.update_client)
        self.client_ac_radio.place(relx=0.805, rely=0.2)

        # Cash A/c RadioButton
        self.cash_ac_radio = ttk.Radiobutton(self.invoice_information_lbl, text="Cash A/c", value="cash",
                                             style="S.TRadiobutton",
                                             variable=self.bill_to_var, command=self.update_client)
        self.cash_ac_radio.place(relx=0.89, rely=0.2)

        # Contact Label
        self.contact_no_lbl = Required_Text(self.invoice_information_lbl, required_text="Contact   ")
        # bg="white")
        self.contact_no_lbl.required_frame.place(relx=0.025, rely=0.5)

        # Contact Entry
        self.contact_no_txt = ttk.Entry(self.invoice_information_lbl, font=("Calibri", 12),
                                        textvariable=self.contact_no_var)
        self.contact_no_txt.place(relx=0.025, rely=0.7, relwidth=0.17)

        # Retail Options Label
        self.retail_options_lbl = Link_Text(self.invoice_information_lbl, link_text="Retail",
                                            link_function=self.retail_options)
        self.retail_options_lbl.link_lbl.place(relx=0.22, rely=0.5)

        # Client Name Label
        self.client_name_lbl = Required_Text(self.invoice_information_lbl, required_text="Client Name   ")
        #  bg="white")
        self.client_name_lbl.required_frame.place(relx=0.32, rely=0.5)

        # Client Name Entry
        self.client_name_txt = AutocompleteCombobox(self.invoice_information_lbl, completevalues=["Select"],
                                                    font=("Calibri", 12))
        self.client_name_txt.place(relx=0.22, rely=0.7, relwidth=0.17)

        self.conn = sqlite3.connect("DB\\Clients.db")
        self.cursor = self.conn.cursor()

        self.cursor.execute("SELECT * FROM CLIENT")

        self.rows = self.cursor.fetchall()

        self.clients = []

        for i in range(0, int(len(self.rows))):
            self.client = self.rows[i][1]
            self.clients.append(self.client)

        self.client_name_txt.config(completevalues=self.clients)

        self.client_name_txt.bind("<<ComboboxSelected>>", self.client_info_add)

        # Client Address Label
        self.client_address_lbl = Required_Text(self.invoice_information_lbl, required_text="Client Address   ")  # bg="white")
        self.client_address_lbl.required_frame.place(relx=0.415, rely=0.5)

        # Client Address Entry
        self.client_address_txt = ttk.Entry(self.invoice_information_lbl, font=("Calibri", 12),
                                            textvariable=self.contact_address_var)
        self.client_address_txt.place(relx=0.415, rely=0.7, relwidth=0.17)

        # Client GSTIN Variable
        self.client_gstin_var = StringVar()

        # Client GSTIN Label
        self.client_gstin_lbl = Required_Text(self.invoice_information_lbl, required_text="Client GSTIN   ")  # bg="white")
        self.client_gstin_lbl.required_frame.place(relx=0.61, rely=0.5)

        # Client GSTIN Entry
        self.client_gstin_txt = ttk.Entry(self.invoice_information_lbl, font=("Calibri", 12),
                                          textvariable=self.client_gstin_var)
        self.client_gstin_txt.place(relx=0.61, rely=0.7, relwidth=0.17)

        # Sold By Label
        self.sold_by_lbl = Required_Text(self.invoice_information_lbl, required_text="Sold By   ")  # bg="white")
        self.sold_by_lbl.required_frame.place(relx=0.805, rely=0.5)

        # Sold By ComboBox
        self.sold_by_txt = ttk.Combobox(self.invoice_information_lbl, font=("Calibri", 12), values=["Select"],
                                        state='readonly')
        self.sold_by_txt.place(relx=0.805, rely=0.7, relwidth=0.17)
        self.sold_by_txt.current(0)

        self.staff_list = ["Select", ]

        self.conn = sqlite3.connect("DB\\Employee.db")
        self.cursor = self.conn.cursor()
        self.cursor.execute("SELECT * FROM STAFF")
        self.rows = self.cursor.fetchall()

        for i in range(0, int(len(self.rows))):
            self.staff_name = self.rows[i][6]
            self.staff_list.append(self.staff_name)

        self.sold_by_txt.config(values=self.staff_list)
        self.sold_by_txt.current(0)

        # Particulars Frame
        self.particulars_lbl = LabelFrame(self.invoice_root, text="Particulars", )  # bg="white")
        self.particulars_lbl.place(relx=0.005, rely=0.22, relwidth=0.99, relheight=0.41)

        # Bar Code / QR Code
        self.bar_var = StringVar()
        self.bar_var.set("bar code")

        # Bar Code RadioButton
        self.bar_code_radio = Radiobutton(self.particulars_lbl, text="Bar Code",
                                          font=("Calibri", 12),  # bg="white",
                                          value="bar code",
                                          variable=self.bar_var)
        self.bar_code_radio.place(relx=0.005, rely=0.08)

        # QR Code RadioButton
        self.item_code_radio = Radiobutton(self.particulars_lbl, text="Item Code",
                                           font=("Calibri", 12),  # bg="white",
                                           value="qr code",
                                           variable=self.bar_var)
        self.item_code_radio.place(relx=0.080, rely=0.08)

        # Bar Code Variable
        self.item_code_var = StringVar()

        # Bar Code Entry
        self.item_code_txt = ttk.Entry(self.particulars_lbl, font=("Calibri", 12), textvariable=self.item_code_var)
        self.item_code_txt.place(relx=0.005, rely=0.3, relwidth=0.155)

        # Item Name Label
        self.item_name_lbl = Required_Text(self.particulars_lbl, required_text="Item Name   ")  # bg="white")
        self.item_name_lbl.required_frame.place(relx=0.165, rely=0)

        # Item Name Entry
        self.item_name_txt = AutocompleteCombobox(self.particulars_lbl, font=("Calibri", 12))
        self.item_name_txt.place(relx=0.165, rely=0.1, relwidth=0.17)

        self.conn = sqlite3.connect("DB\\Items.db")
        self.cursor = self.conn.cursor()

        self.cursor.execute("SELECT * FROM ITEMS")

        self.rows = self.cursor.fetchall()

        self.items_names = []

        for i in range(0, int(len(self.rows))):
            self.items_name = self.rows[i][4]
            self.items_names.append(self.items_name)

        self.item_name_txt.config(completevalues=self.items_names)

        self.item_name_txt.bind("<<ComboboxSelected>>", self.item_info_add)

        # Item Units Label
        self.item_units_lbl = Required_Text(self.particulars_lbl, required_text="Units   ")  # bg="white")
        self.item_units_lbl.required_frame.place(relx=0.340, rely=0)

        # Item Units Entry
        self.item_units_txt = AutocompleteCombobox(self.particulars_lbl, font=("Calibri", 12))
        self.item_units_txt.place(relx=0.340, rely=0.1, relwidth=0.08)

        self.units = []

        self.conn = sqlite3.connect("DB\\Business.db")
        self.cursor = self.conn.cursor()

        self.cursor.execute("SELECT * FROM UNIT")

        self.rows = self.cursor.fetchall()

        for i in self.rows:
            self.units.append(i[1])

        self.item_units_txt.config(completevalues=self.units)

        self.conn = sqlite3.connect("DB\\Items.db")
        self.cursor = self.conn.cursor()

        self.cursor.execute("SELECT * FROM ITEMS")

        self.rows = self.cursor.fetchall()

        self.units_names = []

        for i in range(0, int(len(self.rows))):
            self.units_names.append(self.rows[i][4])

        self.item_name_txt.config(completevalues=self.units_names)

        # Item Quantity Label
        self.item_qty_lbl = Required_Text(self.particulars_lbl, required_text="Item Quantity   ")  # bg="white")
        self.item_qty_lbl.required_frame.place(relx=0.425, rely=0)

        # Item Quantity Variable
        self.item_qty_var = StringVar()

        # Item Quantity Entry
        self.item_qty_txt = ttk.Entry(self.particulars_lbl, font=("Calibri", 12), textvariable=self.item_qty_var)
        self.item_qty_txt.place(relx=0.425, rely=0.1, relwidth=0.08)

        # Item Rate Label
        self.item_rate_lbl = Required_Text(self.particulars_lbl, required_text="Item Rate   ")  # bg="white")
        self.item_rate_lbl.required_frame.place(relx=0.510, rely=0)

        # Item Rate Variable
        self.item_rate_var = StringVar()

        # Item Rate Entry
        self.item_rate_txt = ttk.Entry(self.particulars_lbl, font=("Calibri", 12), textvariable=self.item_rate_var)
        self.item_rate_txt.place(relx=0.510, rely=0.1, relwidth=0.08)

        # Category Label
        self.category_lbl = Required_Text(self.particulars_lbl, required_text="Category   ")  # bg="white")
        self.category_lbl.required_frame.place(relx=0.595, rely=0)

        # Category ComboBox
        self.category_txt = ttk.Combobox(self.particulars_lbl, font=("Calibri", 12), values=["Select"],
                                         state='readonly')
        self.category_txt.place(relx=0.595, rely=0.1, relwidth=0.06)
        self.category_txt.current(0)

        # Connecting To Database
        self.conn = sqlite3.connect("DB\\Business.db")
        self.cursor = self.conn.cursor()

        # Executing Commands In Cursor
        self.cursor.execute("SELECT * FROM CATEGORY")

        # Rows of Data
        self.rows = self.cursor.fetchall()

        # Giving Rows To Values
        self.values = ["Select"]

        # Through For Loop
        for id, category in self.rows:
            # Appending To Values
            self.values.append(category)

        # Giving Values to Category ComboBox
        self.category_txt['values'] = self.values

        # Initializing Items List
        self.items = {}

        # Through For Loop
        for id, category in self.rows:
            # Appending To Items List
            self.items[category] = {}

        # Initializing Categories List
        self.categories = {}

        # Through For Loop
        for id, category in self.rows:
            # Appending To Categories List
            self.categories[category] = {}

        # Sub-Category Label
        self.sub_category_lbl = Required_Text(self.particulars_lbl, required_text="Sub-Category   ")  # bg="white")
        self.sub_category_lbl.required_frame.place(relx=0.660, rely=0)

        # Sub-Category ComboBox
        self.sub_category_txt = ttk.Combobox(self.particulars_lbl, font=("Calibri", 12), values=["Select"])
        self.sub_category_txt.place(relx=0.660, rely=0.1, relwidth=0.06)
        self.sub_category_txt.current(0)

        # Discount Label
        self.discount_lbl = Label(self.particulars_lbl, text="Discount (%)", font=("Calibri", 10), )  # bg="white")
        self.discount_lbl.place(relx=0.725, rely=0)

        # Discount Entry
        self.discount_txt = ttk.Entry(self.particulars_lbl, font=("Calibri", 12))
        self.discount_txt.place(relx=0.725, rely=0.1, relwidth=0.06)
        self.discount_txt.bind("<KeyRelease>", self.discount_amount_info_event)

        # Tax Label
        self.tax_lbl = Label(self.particulars_lbl, text="Tax (%)", font=("Calibri", 10), )  # bg="white")
        self.tax_lbl.place(relx=0.790, rely=0)

        # Tax Entry
        self.tax_txt = ttk.Entry(self.particulars_lbl, font=("Calibri", 12))
        self.tax_txt.place(relx=0.790, rely=0.1, relwidth=0.05)

        # CESS Label
        self.cess_lbl = Label(self.particulars_lbl, text="CESS (%)", font=("Calibri", 10), )  # bg="white")
        self.cess_lbl.place(relx=0.845, rely=0)

        # CESS Entry
        self.cess_txt = ttk.Entry(self.particulars_lbl, font=("Calibri", 12))
        self.cess_txt.place(relx=0.845, rely=0.1, relwidth=0.05)

        # Amount Variable
        self.amount_var = StringVar()

        # Amount Label
        self.amount_lbl = Label(self.particulars_lbl, text="Amount", font=("Calibri", 10))  # bg="white")
        self.amount_lbl.place(relx=0.900, rely=0)

        # Amount Entry
        self.amount_txt = ttk.Entry(self.particulars_lbl, font=("Calibri", 12),
                                    textvariable=self.amount_var)
        self.amount_txt.place(relx=0.900, rely=0.1, relwidth=0.05)

        # Description Label
        self.description_lbl = Label(self.particulars_lbl, text="Description", font=("Calibri", 10), )  # bg="white")
        self.description_lbl.place(relx=0.165, rely=0.20)

        # Description Entry
        self.description_txt = scrolledtext.ScrolledText(self.particulars_lbl, font=("Calibri", 12))
        self.description_txt.place(relx=0.165, rely=0.30, relwidth=0.675, relheight=0.1)

        # Serial number Label
        self.serial_number_lbl = Label(self.particulars_lbl, text="Serial Number",
                                       font=("Calibri", 10), )  # bg="white")
        self.serial_number_lbl.place(relx=0.845, rely=0.20)

        # Serial number Entry
        self.serial_number_txt = ttk.Entry(self.particulars_lbl, font=("Calibri", 12))
        self.serial_number_txt.place(relx=0.845, rely=0.3, relwidth=0.105)

        # Add Particulars Image
        self.add_particular_image = PhotoImage(file="Images\\Add_Button.png")

        # Add Particulars Button
        self.add_particular_btn = ttk.Button(self.particulars_lbl, image=self.add_particular_image,
                                             style="S.TButton",
                                             command=self.add_particular_func)
        self.add_particular_btn.place(relx=0.965, rely=0.1)

        # Particulars Treeview Frame
        self.particulars_treeview_frame = Frame(self.particulars_lbl)
        self.particulars_treeview_frame.place(relx=0, rely=0.45, relwidth=1, relheight=0.55)

        # Particulars Treeview Dictionary
        self.particulars_dict = {
            "No.": {
                "name": "No.",
                "width": "60"
            },
            "Name": {
                "name": "Name",
                "width": "240"
            },
            "Unit": {
                "name": "Unit",
                "width": "70"
            },
            "Rate": {
                "name": "Rate",
                "width": "80"
            },
            "Quantity": {
                "name": "Quantity",
                "width": "80"
            },
            "Discount": {
                "name": "Discount",
                "width": "80"
            },
            "Tax": {
                "name": "Tax",
                "width": "80"
            },
            "CESS": {
                "name": "CESS",
                "width": "80"
            },
            "Amount": {
                "name": "Amount",
                "width": "80"
            },
            "Sub-Category": {
                "name": "Sub-Category",
                "width": "150"
            },
            "Category": {
                "name": "Category",
                "width": "160"
            },
        }

        # Particulars Treeview
        self.particulars_treeview = Custom_treeview(master=self.particulars_treeview_frame,
                                                    command_options=[self.edit_item, self.delete_item],
                                                    command_labels=["Edit", "Delete"], columns=self.particulars_dict,
                                                    )
        self.particulars_treeview.pack(fill=BOTH, expand=1)

        # Payment Frame
        self.payment_lbl = LabelFrame(self.invoice_root, text="Payment", )  # bg="white")
        self.payment_lbl.place(relx=0.59, rely=0.635, relwidth=0.18, relheight=0.360)

        # Payment Date Label
        self.payment_date_lbl = Label(self.payment_lbl, text="Date",  # bg="white",
                                      font=("Calibri", 11))
        self.payment_date_lbl.place(relx=0, rely=0.01)

        # Payment Date Entry
        self.payment_date_txt = DateEntry(self.payment_lbl, font=("Calibri", 11))
        self.payment_date_txt.place(relx=0.3, rely=0.01, relwidth=0.6)

        # Payment Mode Label
        self.payment_mode_lbl = Label(self.payment_lbl, text="Mode",  # bg="white",
                                      font=("Calibri", 11))
        self.payment_mode_lbl.place(relx=0, rely=0.2)

        # Payment Mode ComboBox
        self.payment_mode_txt = ttk.Combobox(self.payment_lbl, font=("Calibri", 11), values=["Select", "Cash",
                                                                                             "Cheque", "Card",
                                                                                             "Demand Draft",
                                                                                             "Mobile Wallet",
                                                                                             "Bank Transfer"],
                                             state='readonly', style="TCombobox")
        self.payment_mode_txt.current(0)
        self.payment_mode_txt.place(relx=0.3, rely=0.2, relwidth=0.6)

        self.payment_mode_txt.bind("<<ComboboxSelected>>", self.update_payment_info_event)

        # Txn. Id Label
        self.txn_id_lbl = Label(self.payment_lbl, text="Txn. Id",  # bg="white",
                                font=("Calibri", 11))
        self.txn_id_lbl.place(relx=0, rely=0.4)

        # Txn. Id Entry
        self.txn_id_txt = ttk.Entry(self.payment_lbl, font=("Calibri", 11), state='disabled')
        self.txn_id_txt.place(relx=0.3, rely=0.4, relwidth=0.6)

        # Payment Amount Label
        self.payment_amount_lbl = Label(self.payment_lbl, text="Amount",  # bg="white",
                                        font=("Calibri", 11))
        self.payment_amount_lbl.place(relx=0, rely=0.6)

        # Payment Amount Variable
        self.payment_amount_var = StringVar()

        # Payment Amount Entry
        self.payment_amount_txt = ttk.Entry(self.payment_lbl, text="", font=("Calibri", 11),
                                            textvariable=self.payment_amount_var)
        self.payment_amount_txt.place(relx=0.3, rely=0.6, relwidth=0.6)

        # Payment Balance Label
        self.payment_balance_lbl = Label(self.payment_lbl, text="Balance",  # bg="white",
                                         font=("Calibri", 11))
        self.payment_balance_lbl.place(relx=0, rely=0.8)

        # Payment Balance variable
        self.payment_balance_var = StringVar()

        # Payment Amount Entry
        self.payment_balance_txt = ttk.Entry(self.payment_lbl, font=("Calibri", 11), state='disabled',
                                             textvariable=self.payment_balance_var)
        self.payment_balance_txt.place(relx=0.3, rely=0.8, relwidth=0.6)

        # Variable for Discount CheckButton
        self.discount_var = IntVar()

        # Discount CheckButton
        self.discount_chkbox = ttk.Checkbutton(self.invoice_root, text="Apply Discount To All",
                                               variable=self.discount_var, onvalue=1, offvalue=0, takefocus=False,
                                               style='Red.TCheckbutton', command=self.add_discount)

        # Packing Discount CheckButton
        self.discount_chkbox.place(relx=0.11, rely=0.77)

        # Variable for Discount CheckButton
        self.shipping_var = IntVar()
        self.shipping_var.set(1)

        # Shipping CheckButton
        self.shipping_chkbox = ttk.Checkbutton(self.invoice_root, text="Add Shipping",
                                               variable=self.shipping_var, onvalue=1, offvalue=0, takefocus=False,
                                               style='Red.TCheckbutton', command=self.add_shipping)
        self.add_shipping()

        # Packing Shipping CheckButton
        self.shipping_chkbox.place(relx=0.33, rely=0.77)

        # Delivery Terms Frame
        self.delivery_terms_lbl = LabelFrame(self.invoice_root, text="Delivery Terms", )  # bg="white")
        self.delivery_terms_lbl.place(relx=0.25, rely=0.88, relwidth=0.22, relheight=0.1)

        # Delivery Terms Entry
        self.delivery_terms_txt = scrolledtext.ScrolledText(self.delivery_terms_lbl, font=("Calibri", 12))
        self.delivery_terms_txt.pack(fill=BOTH, expand=1)

        # Remarks Frame
        self.remarks_lbl = LabelFrame(self.invoice_root, text="Remarks(Private Use)", )  # bg="white")
        self.remarks_lbl.place(relx=0.01, rely=0.88, relwidth=0.22, relheight=0.1)

        # Remarks Entry
        self.remarks_txt = scrolledtext.ScrolledText(self.remarks_lbl, font=("Calibri", 12))
        self.remarks_txt.pack(fill=BOTH, expand=1)
        self.invoice_root.update_idletasks()

        # Total Amount Frame
        self.total_amount_frame = LabelFrame(self.invoice_root, text="Total Amount", )  # bg="white")
        self.total_amount_frame.place(relx=0.78, rely=0.635, relwidth=0.21, relheight=0.2)

        # Total Price Variable
        self.totalPrice = StringVar()

        # Discount Price Variable
        self.discountPrice = StringVar()

        # Shipping Charges Variable
        self.shippingCharges = StringVar()

        # Subtotal Amount Variable
        self.subTotalPrice = StringVar()

        # Total Amount Label
        self.total_amount_lbl = Label(self.total_amount_frame, text="SubTotal:-", font=("Calibri", 13), )  # bg="white")
        self.total_amount_lbl.place(relx=0, rely=0)

        # Total Amount TXT
        self.total_amount_txt = Label(self.total_amount_frame, text="", textvariable=self.subTotalPrice,  # bg="white",
                                      font=("Calibri", 13))
        self.total_amount_txt.place(relx=0.3, rely=0)

        # Shipping Charges Label
        self.shipping_charges_lbl = Label(self.total_amount_frame, text="Shipping:-",
                                          font=("Calibri", 13), )  # bg="white")
        self.shipping_charges_lbl.place(relx=0, rely=0.25)

        # Shipping Charges TXT
        self.shippingCharges_txt = Label(self.total_amount_frame, text="", textvariable=self.shippingCharges,
                                         # bg="white",
                                         font=("Calibri", 13))
        self.shippingCharges_txt.place(relx=0.3, rely=0.25)

        # Discount Amount Label
        self.discount_amount_lbl = Label(self.total_amount_frame, text="Discount:-",
                                         font=("Calibri", 13), )  # bg="white")
        self.discount_amount_lbl.place(relx=0, rely=0.50)

        # Discount Amount TXT
        self.discount_amount_txt = Label(self.total_amount_frame, text="", textvariable=self.discountPrice,
                                         # bg="white",
                                         font=("Calibri", 13))
        self.discount_amount_txt.place(relx=0.3, rely=0.50)

        # Total Amount Label
        self.total_amount_lbl = Label(self.total_amount_frame, text="Total:-", font=("Calibri", 13), )  # bg="white")
        self.total_amount_lbl.place(relx=0, rely=0.75)

        # Total Amount TXT
        self.total_amount_txt = Label(self.total_amount_frame, text="", textvariable=self.totalPrice,  # bg="white",
                                      font=("Calibri", 13))
        self.total_amount_txt.place(relx=0.3, rely=0.75)

        # Save Button
        self.save_btn = ttk.Button(self.invoice_root, text="Save", style="S.TButton", command=self.save_operation)
        self.save_btn.place(relx=0.8, rely=0.88)

        # Save & Print Button
        self.save_print_btn = ttk.Button(self.invoice_root, text="Save & Print", style="S.TButton",
                                         command=self.print_operation)
        self.save_print_btn.place(relx=0.9, rely=0.88)

        self.update_client()

    def discount_amount_info_event(self, event):
        self.discount = float(self.discount_txt.get())
        self.rate = int(self.item_rate_txt.get())
        self.quantity = int(self.item_qty_txt.get())

        self.final_rate = self.rate * self.quantity
        self.final_discount = (self.discount * self.final_rate) / 100
        self.final_amount = self.amount_var.set(self.final_rate - self.final_discount)

    def update_payment_info_event(self, event):
        if self.payment_mode_txt.get() == "Cheque" or \
                self.payment_mode_txt.get() == "Mobile Wallet" or self.payment_mode_txt.get() == "Bank Transfer":
            self.txn_id_lbl.config(text="Txn. Id")
            self.txn_id_txt.config(state=NORMAL)
        elif self.payment_mode_txt.get() == "Demand Draft":
            self.txn_id_lbl.config(text="DD No.")
            self.txn_id_txt.config(state=NORMAL)
        elif self.payment_mode_txt.get() == "Card":
            self.txn_id_lbl.config(text="Last 4 Digits")
            self.txn_id_txt.config(state=NORMAL)
        else:
            self.txn_id_lbl.config(text="Txn. Id")
            self.txn_id_txt.config(state=DISABLED)

    def edit_item(self):
        self.edit_window = Toplevel()
        self.edit_window.geometry("600x500+500+400")

    def delete_item(self):
        self.particulars_treeview.custom_treeview.delete(self.particulars_treeview.cursor_row)
        self.contents = self.particulars_treeview.contents.values()
        self.contents1 = list(self.contents)
        for i in self.categories.keys():
            for j in self.items[i].keys():
                if self.items[i][j][0] == self.contents1[2][1]:
                    self.items[i].pop(j)
                    self.update_total_price()
                    break

    # Adding Items Info
    def item_info_add(self, event):
        self.conn = sqlite3.connect("DB\\Items.db")
        self.cursor = self.conn.cursor()

        self.cursor.execute("SELECT * FROM ITEMS where Item_Name = ?", (self.item_name_txt.get(),))
        self.rows = self.cursor.fetchone()

        self.category_name = self.rows[1]
        self.sub_category_name = self.rows[2]
        self.item_code_name = self.rows[3]
        self.item_rate_name = self.rows[5]
        self.item_qty_name = self.rows[6]

        self.category_txt.set(self.category_name)
        self.sub_category_txt.set(self.sub_category_name)
        self.item_code_var.set(self.item_code_name)
        self.item_rate_var.set(self.item_rate_name)
        self.item_qty_var.set("1")

    def update_client(self):
        if self.bill_to_var.get() == "cash":
            self.client_name_var = StringVar()
            self.client_name_txt = ttk.Entry(self.invoice_information_lbl, font=("Calibri", 12),
                                             textvariable=self.client_name_var)
            self.client_name_txt.place(relx=0.22, rely=0.7, relwidth=0.17)
            self.client_name_var.set("CASH")
            self.payment_balance_var.set(0)
        else:
            # Client Name Entry
            self.client_name_txt = AutocompleteCombobox(self.invoice_information_lbl, completevalues=["Select"],
                                                        font=("Calibri", 12))
            self.client_name_txt.place(relx=0.22, rely=0.7, relwidth=0.17)

            self.conn = sqlite3.connect("DB\\Clients.db")
            self.cursor = self.conn.cursor()

            self.cursor.execute("SELECT * FROM CLIENT")

            self.rows = self.cursor.fetchall()

            self.clients = []

            for i in range(0, int(len(self.rows))):
                self.client = self.rows[i][1]
                self.clients.append(self.client)

            self.client_name_txt.config(completevalues=self.clients)

            self.client_name_txt.bind("<<ComboboxSelected>>", self.client_info_add)

        self.invoice_type_txt.current(2)

    # Adding Client Info
    def client_info_add(self, event):
        self.conn = sqlite3.connect("DB\\Clients.db")
        self.cursor = self.conn.cursor()

        self.cursor.execute("SELECT * FROM CLIENT where Full_Name = ?", (self.client_name_txt.get(),))
        self.rows = self.cursor.fetchone()

        self.client_contact = self.rows[5]
        self.client_address = self.rows[8]
        self.client_balance = self.rows[22]

        self.contact_no_var.set(self.client_contact)
        self.contact_address_var.set(self.client_address)
        self.payment_balance_var.set(self.client_balance)

    # Retail Options Window Function
    def retail_options(self, event):

        # Retail Options Window
        self.retail_options_window = Toplevel()
        self.retail_options_window.overrideredirect(True)
        self.retail_options_window.geometry("500x300+400+200")

        # Dict for Retail Option Values
        self.retail_options = ["Retail", "Agent", "Wholesale"]

        # Retail Frame
        self.retail_frame = ttk.Frame(self.retail_options_window, style="Retail.TFrame")
        self.retail_frame.pack(fill=BOTH, side=TOP, anchor=N, expand=1)

        # Retail Button
        self.retail_lbl = ttk.Button(self.retail_frame, text=self.retail_options[0], command=self.retail_operation
                                     , style="S.TButton")
        self.retail_lbl.pack(fill=BOTH, side=TOP, expand=1)

        # Agent Frame
        self.agent_frame = ttk.Frame(self.retail_options_window, style="Retail.TFrame")
        self.agent_frame.pack(fill=BOTH, side=TOP, anchor=N, expand=1)

        # Agent Button
        self.agent_lbl = ttk.Button(self.retail_frame, text=self.retail_options[1], command=self.agent_operation
                                    , style="S.TButton")
        self.agent_lbl.pack(fill=BOTH, side=TOP, expand=1)

        # Wholesale Frame
        self.wholesale_frame = ttk.Frame(self.retail_options_window, style="Retail.TFrame")
        self.wholesale_frame.pack(fill=BOTH, side=TOP, anchor=N, expand=1)

        # Wholesale Button
        self.wholesale_lbl = ttk.Button(self.retail_frame, text=self.retail_options[2], command=self.wholesale_operation
                                        , style="S.TButton")
        self.wholesale_lbl.pack(fill=BOTH, side=TOP, expand=1)

    # Retail Button Function
    def retail_operation(self):
        self.retail_options_lbl.link_lbl.config(text="Retail")
        self.retail_options_window.destroy()

    # Agent Button Function
    def agent_operation(self):
        self.retail_options_lbl.link_lbl.config(text="Agent")
        self.retail_options_window.destroy()

    # Wholesale Button Function
    def wholesale_operation(self):
        self.retail_options_lbl.link_lbl.config(text="Wholesale")
        self.retail_options_window.destroy()

    # Adding Discount Option
    def add_discount(self):
        # IF ELSE Statement For Adding And Removing Discount Entries According to CheckButton
        if self.discount_var.get() == 1:
            self.discount_charges_txt = ttk.Entry(self.invoice_root, font=("Calibri", 12))
            self.discount_charges_txt.place(relx=0.25, rely=0.77, relwidth=0.05)
            self.discount_charges_txt.bind("<KeyRelease>", self.update_total_price_event)
        else:
            self.discount_charges_txt.destroy()

    def update_total_price_event(self, event):
        self.update_total_price()

    # Saving Bills
    def save_operation(self):
        self.conn = sqlite3.connect("DB\\Invoices.db")
        self.cursor = self.conn.cursor()

        self.invoice_n = re.findall(r"[^\W\d_]+|\d+", self.invoice_no)
        self.invoice = int(self.invoice_n[1])
        self.invoice += 1

        self.cwd = os.getcwd()

        self.file_path = f"{self.cwd}\\DB\\INVOICES"

        if not os.path.exists(self.file_path):
            os.makedirs(self.file_path)

        # self.conn.execute("DROP TABLE INVOICES")

        # Preparing Table For Invoices
        self.query = """CREATE TABLE IF NOT EXISTS INVOICES(
                        id  INTEGER PRIMARY KEY autoincrement,
                        Invoice_Type VARCHAR(500) NOT NULL,
                        Invoice_No VARCHAR(500) NOT NULL,
                        Invoice_Date DATE NOT NULL,
                        Sold_By VARCHAR(500) NOT NULL,
                        Client_Contact INT(10) NOT NULL,
                        Client_Name VARCHAR(500) NOT NULL,
                        Client_Address VARCHAR(5000) NOT NULL,
                        Discount INTEGER NOT NULL,
                        Shipping INTEGER NOT NULL,
                        SubTotal INTEGER NOT NULL,
                        Total INTEGER NOT NULL,
                        Payment_Date DATE NOT NULL,
                        Payment_Mode VARCHAR(500) NOT NULL,
                        Payment_No VARCHAR(500) NOT NULL,
                        Payment_Amount VARCHAR(500) NOT NULL,
                        Client_Balance INTEGER NOT NULL,
                        Remarks VARCHAR(5000) NOT NULL,
                        Delivery_Terms VARCHAR(5000) NOT NULL)
                     """
        self.save_bill_query = """INSERT INTO INVOICES(Invoice_Type, Invoice_No, Invoice_Date,
                                                        Sold_By, Client_Contact, Client_Name,
                                                        Client_Address, Discount, Shipping,
                                                        SubTotal, Total, Payment_Date, Payment_Mode,
                                                        Payment_No, Payment_Amount, Client_Balance,
                                                        Remarks, Delivery_Terms) 
                                                        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"""
        self.save_bill_values = (self.invoice_type_txt.get(), self.invoice_no_txt.get(),
                                 self.invoice_date_txt.get(), self.sold_by_txt.get(),
                                 self.contact_no_txt.get(), self.client_name_txt.get(),
                                 self.client_address_txt.get(), self.discountPrice.get(),
                                 self.shippingCharges.get(), self.subTotalPrice.get(),
                                 self.totalPrice.get(), self.payment_date_txt.get(),
                                 self.payment_mode_txt.get(), self.txn_id_txt.get(),
                                 self.payment_amount_txt.get(), self.payment_balance_txt.get(),
                                 self.remarks_txt.get("1.0", END), self.delivery_terms_txt.get("1.0", END),
                                 )
        self.cursor.execute(self.query)
        self.cursor.execute(self.save_bill_query, self.save_bill_values)
        self.conn.commit()
        self.conn.close()
        self.save_items = json.dumps(self.items, indent=len(self.items))
        with open(f"Details\\{self.client_name_txt.get()}{self.invoice_no_txt.get()}.json", "w") as outfile:
            outfile.write(self.save_items)
        self.save_excel()
        self.conn = sqlite3.connect("DB\\Business.db")
        self.cursor = self.conn.cursor()
        self.sql_string = '''UPDATE ID
                             SET Id = ?
                             WHERE Category = ?'''
        self.sql_string_query = (self.sql_string, (self.invoice, "Invoice"))
        self.conn.commit()
        self.conn.close()
        self.ans = tmsg.showinfo("Success", "Data Saved! Click Ok To Exit")
        self.invoice_root.destroy()

    # Saving Excel Function
    def save_excel(self):

        # Getting Details From Database
        self.conn = sqlite3.connect("DB\\Invoices.db")
        self.cursor = self.conn.cursor()
        self.cursor.execute("SELECT * FROM INVOICES where Invoice_No = ?", (self.invoice_no_txt.get(),))
        self.rows = self.cursor.fetchone()

        # Excel File Starting
        self.xlWorkBook = load_workbook(f"Images\\Workbook_2.xlsx")
        self.xlSheet = self.xlWorkBook.active
        del self.xlWorkBook['Invoice']._images[0]
        self.xlSheet.title = "Invoice"
        self.headerxl = ["No.", "Name", "Rate", "Quantity", "Price", "Total Price:- ",  # + totalPrice.get()
                         str(date.today())]
        self.delivery_date = str(date.today() + timedelta(4))
        self.col = 0
        self.row = 0
        self.img = openpyxl.drawing.image.Image('Images\\Logo.png')
        self.img.anchor = "F1"
        self.img.width = 213.165312
        self.img.height = 91.0866144
        self.xlSheet.add_image(self.img)
        self.xlSheet.cell(row=self.row + 9, column=self.col + 1, value=f"Customer Name:- {self.rows[7]}")
        self.xlSheet.cell(row=self.row + 10, column=self.col + 1, value=f"Customer Contact No:- {self.rows[5]}")
        self.xlSheet.cell(row=self.row + 12, column=self.col + 1, value=self.rows[8])
        self.xlSheet.cell(row=self.row + 9, column=self.col + 5, value=self.headerxl[6])
        self.xlSheet.cell(row=self.row + 6, column=self.col + 7, value=self.headerxl[6])
        self.xlSheet.cell(row=self.row + 6, column=self.col + 5, value=self.rows[2])
        self.xlSheet.cell(row=self.row + 11, column=self.col + 5, value=str(self.delivery_date))
        self.xlSheet.cell(row=self.row + 32, column=self.col + 7, value=self.rows[11])
        self.xlSheet.print_area = 'A1:G42'
        self.xlSheet.print_options.verticalCentered = True
        self.xlSheet.print_options.horizontalCentered = True
        self.row1 = 17

        for i in self.categories:
            for j in self.items[i].keys():
                self.lis = self.items[i][j]
                self.names1 = self.lis[0]
                self.rates1 = self.lis[1]
                self.units1 = self.lis[7]
                self.quantitys1 = self.lis[2]
                self.prices1 = self.lis[3]
                self.xlSheet.cell(row=self.row1, column=self.col + 1, value=self.names1)
                self.xlSheet.cell(row=self.row1, column=self.col + 5, value=self.quantitys1)
                self.xlSheet.cell(row=self.row1, column=self.col + 6, value=self.rates1)
                self.xlSheet.cell(row=self.row1, column=self.col + 7, value=self.units1)
                self.xlSheet.cell(row=self.row1, column=self.col + 8, value=self.prices1)
                self.row1 = self.row1 + 1

        self.xlWorkBook.save(f"Bill Records\\{self.rows[7]}{self.rows[2]}.xlsx")
        self.xlWorkBook.close()

    def print_excel(self):

        # Getting Details From Database
        self.conn = sqlite3.connect("DB\\Invoices.db")
        self.cursor = self.conn.cursor()
        self.cursor.execute("SELECT * FROM INVOICES where Invoice_No = ?", (self.invoice_no_txt.get(),))
        self.rows = self.cursor.fetchone()

        # Saving As PNG To Preview
        # Path to original excel file
        self.WB_PATH = f"{os.getcwd()}\\Bill Records\\{self.rows[7]}{self.rows[2]}.xlsx"
        # PDF path when saving
        self.PATH_TO_PDF = f"{os.getcwd()}\\Bill Records\\{self.rows[7]}{self.rows[2]}.pdf"

        self.excel = win32com.client.Dispatch("Excel.Application")

        self.excel.Visible = False

        try:
            # Open
            self.wb = self.excel.Workbooks.Open(self.WB_PATH)

            # Specify the sheet you want to save by index. 1 is the first (leftmost) sheet.
            self.ws_index_list = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]
            self.wb.WorkSheets(self.ws_index_list[0]).Select()

            # Save
            self.wb.ActiveSheet.ExportAsFixedFormat(0, self.PATH_TO_PDF)
        except com_error as e:
            print(f"failed. due to {e}")
            return
        else:
            self.wb.Close()
            self.excel.Quit()

        self.f = f"{os.getcwd()}\\Bill Records\\{self.rows[7]}{self.rows[2]}.pdf"
        self.newfilename = ""
        try:
            with(wand.image.Image(filename=self.f, resolution=120)) as source:
                for i, image in enumerate(source.sequence):
                    self.newfilename = self.f[:-4] + str(i + 1) + '.png'
                    wand.image.Image(image).save(filename=self.newfilename)
        except wand.exceptions.DelegateError as e:
            print(e)

    # Saving & Printing Bills
    def print_operation(self):
        self.ans = tmsg.askquestion("Are You Sure?", "Are You Sure To Print Bill?")
        if self.ans == "yes":
            # Loading Start
            self.loading_start = time.time()
            self.loading_screen = Tk()
            self.loading_screen.overrideredirect(True)
            self.loading_screen.focus_set()
            self.loading_bar = ttk.Progressbar(self.loading_screen, orient=HORIZONTAL, length=200,
                                               mode="determinate", maximum=100, value=20)
            self.loading_bar.pack()
            # Save Function
            self.save_operation()
            self.loading_bar.step(50)
            # Print Function
            self.print_excel()
            self.invoice_root.update_idletasks()
            self.invoice = self.invoice_no_txt.get()
            # Destroying Root Window
            self.invoice_root.destroy()
            del self.invoice_root
            self.loading_bar.step(30)
            self.loading_screen.destroy()

            # Loading End
            self.loading_end = time.time()

            # Print Preview Window
            self.print_preview_window = Tk()
            self.print_preview_window['bg'] = "white"

            self.print_frame = Frame(self.print_preview_window, bg="white")
            self.print_frame.place(relx=0.3, rely=0, relwidth=0.7, relheight=1)

            self.print_btn = ttk.Button(self.print_preview_window, text="Print", style="S.TButton",
                                        command=self.print_button)
            self.print_btn.place(relx=0, rely=0.1)
            try:
                self.print_image = PhotoImage(file=self.newfilename)
                self.print_page = Label(self.print_frame, image=self.print_image)
                self.print_page.pack(anchor=CENTER, fill=BOTH)
            except Exception as e:
                print(e)

    # Print Button
    def print_button(self):
        # Getting Details From Database
        self.conn = sqlite3.connect("DB\\Invoices.db")
        self.cursor = self.conn.cursor()
        self.cursor.execute("SELECT * FROM INVOICES where Invoice_No = ?", (self.invoice,))
        self.rows = self.cursor.fetchone()

        # Printing Excel File
        # Path to original excel file
        self.WB_PATH = f"{os.getcwd()}\\Bill Records\\{self.rows[7]}{self.rows[2]}.xlsx"

        self.excel = win32com.client.Dispatch("Excel.Application")

        self.excel.Visible = False
        try:
            # Open
            self.wb = self.excel.Workbooks.Open(self.WB_PATH)

            # Specify the sheet you want to save by index. 1 is the first (leftmost) sheet.
            self.ws_index_list = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]
            self.wb.WorkSheets(self.ws_index_list[0]).Select()

            # Print
            self.wb.ActiveSheet.PrintOut()

            # Deleting All Files
            os.remove(self.WB_PATH)
            os.remove(self.newfilename)
            os.remove(self.f)

        except com_error as e:
            print(f"failed. due to {e}")
            return
        else:
            self.print_preview_window.destroy()
            self.wb.Save(True)
            self.wb.Close()
            self.excel.Quit()
            del self.print_preview_window
            del self.wb
            del self.excel

    def update_total_price(self):
        self.subPrice = 0
        self.totalprice = 0
        self.discountprice = 0
        self.shippingprice = 0

        for i in self.categories.keys():
            for j in self.items[i].keys():
                self.subPrice += int(self.items[i][j][3])
                self.discountprice += int(int(self.items[i][j][3]) - float(self.items[i][j][10]))
        self.totalprice += self.subPrice
        self.totalprice -= self.discountprice
        if self.subPrice == 0:
            self.subTotalPrice.set("")
        else:
            self.subTotalPrice.set("Rs." + str(self.subPrice) + " /-")
            self.totalPrice.set("Rs." + str(self.totalprice) + "/-")
            self.payment_amount_var.set(str(self.totalprice))
        if self.discount_var.get() == 1 and self.discount_charges_txt.get() != "":

            self.discountprice += int(self.discount_charges_txt.get())
            self.discountPrice.set("Rs. " + str(self.discountprice) + "/-")

            self.totalprice -= int(self.discountprice)
            self.totalPrice.set("Rs." + str(self.totalprice) + "/-")
        else:
            self.discountPrice.set("Rs. " + str(self.discountprice) + "/-")
        if self.shipping_var.get() == 1:
            if self.shipping_charges_sel and self.shipping_charges_txt:
                if self.shipping_charges_sel.get() != "Select" and self.shipping_charges_txt.get() == "":
                    self.shippingprice = int(self.shipping_charges_sel.get())
                    self.shippingCharges.set("Rs. " + str(self.shippingprice) + "/-")
                    self.totalprice += self.shippingprice
                    self.totalPrice.set("Rs." + str(self.totalprice) + "/-")
                elif self.shipping_charges_txt.get() != "":
                    self.shipping_charges_sel.set("Select")
                    self.shippingprice = int(self.shipping_charges_txt.get())
                    self.shippingCharges.set("Rs. " + str(self.shippingprice) + "/-")
                    self.totalprice += self.shippingprice
                    self.totalPrice.set("Rs." + str(self.totalprice) + "/-")
                elif self.shipping_charges_sel.get() == "Select" and self.shipping_charges_txt.get() == "":
                    self.shippingprice = 0
                    self.shippingCharges.set("Rs. " + str(self.shippingprice) + "/-")
                    self.totalprice += self.shippingprice
                    self.totalPrice.set("Rs." + str(self.totalprice) + "/-")
        else:
            self.shippingCharges.set("Rs. " + str(self.shippingprice) + "/-")
        self.payment_amount_var.set(str(self.totalprice))

    # Getting Object Entries From Outside
    def get_object(self, object):
        get_objec = object
        try:
            get_objec.get()
        except Exception as e:
            print(e)
        return get_objec

    # Setting Object Entries From Outside
    def set_object(self, object, value):
        set_objec = object
        try:
            set_objec.set(value)
        except Exception as e:
            print(e)
        return set_objec

    # Adding Shipping Entries
    def add_shipping(self):

        # IF ELSE Statement For Adding And Removing Shipping Entries According to CheckButton
        if self.shipping_var.get() == 1:
            self.shipping_charges_txt = ttk.Entry(self.invoice_root, font=("Calibri", 12))
            self.shipping_charges_txt.place(relx=0.43, rely=0.77, relwidth=0.05)
            self.shipping_charges_sel = ttk.Combobox(self.invoice_root, values=["Select", "10", "15", "25", "50"])
            self.shipping_charges_sel.place(relx=0.49, rely=0.77, relwidth=0.05)
            self.shipping_charges_sel.current(0)
            self.shipping_charges_txt.bind("<KeyRelease>", self.update_total_price_event)
            self.shipping_charges_sel.bind("<<ComboboxSelected>>", self.update_total_price_event)
        else:
            self.shipping_charges_txt.destroy()
            self.shipping_charges_sel.destroy()

    # Adding Particulars To Particulars Treeview
    def add_particular_func(self):
        if self.category_txt.get() == "Select" or self.item_name_txt.get() == "" or self.item_rate_txt.get() == "" or \
                self.item_units_txt.get() == "":
            tmsg.showerror("Error", "Please Fill All Required Fields.", parent=self.invoice_root)
        else:
            # Variables For Getting Row Values
            self.name = self.item_name_txt.get()
            self.rate = self.item_rate_txt.get()
            self.qty = self.item_qty_txt.get()
            self.item_code = self.item_code_txt.get()
            self.sub_category = self.sub_category_txt.get()
            self.category = self.category_txt.get()
            self.discount = self.discount_txt.get()
            self.tax = self.tax_txt.get()
            self.cess = self.cess_txt.get()
            self.amount = self.amount_txt.get()
            self.unit = self.item_units_txt.get()

            # IF Statement For Checking If Name Already Exist in Order
            if self.name in self.items[self.category].keys():
                tmsg.showinfo("Error", "Item already exist in your order")
                return
            # IF Statement For Checking If Quantity Is Digit in Order
            if not self.qty.isdigit():
                tmsg.showinfo("Error", "Please Enter Valid Quantity")
                return

            # Getting Items in List Variable
            self.lis = [self.name, self.rate, self.qty, str(int(self.rate) * int(self.qty)), self.item_code,
                        self.sub_category,
                        self.category, self.discount, self.tax, self.cess, self.amount, self.unit]
            self.items[self.category][self.name] = self.lis

            # Getting The Last Row Of Particulars Treeview
            self.last_row = len(self.particulars_treeview.custom_treeview.get_children("")) + 1

            # IF ELSE Statement For Getting Last Row And Inserting It To Particulars Treeview
            if self.last_row != 0:
                if self.last_row % 2 == 0:
                    self.particulars_treeview.custom_treeview.insert('', END,
                                                                     values=(
                                                                         self.last_row, self.name, self.unit,
                                                                         self.rate, self.qty, self.discount,
                                                                         self.tax, self.cess, self.amount,
                                                                         self.sub_category, self.category
                                                                     ),
                                                                     tags=('evenrow',))
                else:
                    self.particulars_treeview.custom_treeview.insert('', END,
                                                                     values=(
                                                                         self.last_row, self.name, self.unit,
                                                                         self.rate, self.qty, self.discount,
                                                                         self.tax, self.cess, self.amount,
                                                                         self.sub_category, self.category
                                                                     ),
                                                                     tags=('oddrow',))
                self.update_total_price()
            else:
                self.particulars_treeview.custom_treeview.insert('', END,
                                                                 values=(
                                                                     1, self.name, self.unit,
                                                                     self.rate, self.qty, self.discount,
                                                                     self.tax, self.cess, self.amount,
                                                                     self.sub_category, self.category
                                                                 ),
                                                                 tags=('oddrow',))
                self.update_total_price()

    # Item Menu Function
    def select_from_menu(self):

        # Connecting to Database
        self.conn = sqlite3.connect('DB\\Items.db')
        self.cursor = self.conn.cursor()

        # Item Menu Window
        self.item_menu = Toplevel()

        # Style For TTK
        self.style = ttk.Style(self.item_menu)

        # Using Theme AWLIGHT
        self.style.theme_use('awlight')

        # Treeview Style
        self.style.configure("T.Treeview", background="black", fieldbackground="white", foreground="black")

        # Treeview Heading Style
        self.style.configure("T.Treeview.Heading", background="#26e881", foreground="white", rowheight=35,
                             font=("Calibri", 15))

        # Functions For Item Menu Window
        self.item_menu.title("Product Menu")
        self.item_menu.geometry("600x600+100+12")
        self.item_menu.attributes('-topmost', 1)
        self.item_menu.attributes('-toolwindow', 1)
        self.item_menu.focus_force()

        # Treeview For Items
        self.item_treeview = ttk.Treeview(self.item_menu, style="T.Treeview",
                                          columns=(
                                              "No.", "Name", "Rate",
                                              "Quantity", "Item Code", "Sub-Category", "Category"), )
        # height=len(self.rows))

        # Vertical Scrollbar
        self.vsb = Scrollbar(self.item_treeview,
                             orient="vertical",
                             command=self.item_treeview.yview
                             )

        # Horizontal Scrollbar
        self.hsb = Scrollbar(self.item_treeview,
                             orient="horizontal",
                             command=self.item_treeview.xview
                             )

        # Scroll Command For Treeview
        self.item_treeview['yscrollcommand'] = self.vsb.set
        self.item_treeview['xscrollcommand'] = self.hsb.set

        # Heading For Treeview
        self.item_treeview.heading("No.", text="No.")
        self.item_treeview.heading("Name", text="Name")
        self.item_treeview.heading("Rate", text="Rate")
        self.item_treeview.heading("Quantity", text="Quantity")
        self.item_treeview.heading("Item Code", text="Item Code")
        self.item_treeview.heading("Sub-Category", text="Sub-Category")
        self.item_treeview.heading("Category", text="Category")
        self.item_treeview["displaycolumns"] = ("No.", "Name", "Rate",
                                                "Quantity", "Item Code", "Sub-Category", "Category")
        self.item_treeview["show"] = "headings"

        # Columns For Treeview
        self.item_treeview.column("No.", width=150, anchor='center')
        self.item_treeview.column("Name", width=150, anchor='center')
        self.item_treeview.column("Rate", width=150, anchor='center')
        self.item_treeview.column("Quantity", width=150, anchor='center')
        self.item_treeview.column("Item Code", width=150, anchor='center')
        self.item_treeview.column("Sub-Category", width=150, anchor='center')
        self.item_treeview.column("Category", width=150, anchor='center')

        # Tags For Treeview
        self.item_treeview.tag_configure('oddrow', background="white")
        self.item_treeview.tag_configure('evenrow', background="#c5dbbf")

        # Creating Table In Database
        sql = '''CREATE TABLE IF NOT EXISTS ITEMS(
           id  INTEGER PRIMARY KEY,
           Category VARCHAR(50) NOT NULL,
           Sub_Category VARCHAR(50) NOT NULL,
           Item_Code VARCHAR(500) NOT NULL,
           Item_Name VARCHAR(500) NOT NULL,
           Item_Rate INTEGER NOT NULL,
           Item_Qty INTEGER NOT NULL)'''

        # Fake Data For Database
        values = [('EXAMPLE', '', 'a', 'a', '500', '5'),
                  ('EXAMPLE', '', 'b', 'b', '515', '7'),
                  ('EXAMPLE', '', 'c', 'c', '800', '6'),
                  ('EXAMPLE', '', 'd', 'd', '545', '17'),
                  ('EXAMPLE', '', 'e', 'e', '657', '21'),
                  ('EXAMPLE', '', 'f', 'f', '491', '56'),
                  ('EXAMPLE', '', 'g', 'g', '561', '87'),
                  ('EXAMPLE', '', 'g', 'name', '345', '78')]
        # Commands For Inserting Fake Data In Items Treeview
        '''self.cursor.executemany("""INSERT INTO ITEMS(Category, Sub_Category, Item_Code,
         Item_Name, Item_Rate, Item_Qty) VALUES(?,?,?,?,?,?)""", values)'''
        # For Deleting Items From Treeview For Delete Button TESTING
        '''self.cursor.execute("DELETE FROM ITEMS WHERE Category = ?", 'a')
        self.cursor.execute("DELETE FROM ITEMS WHERE Category = ?", 'g')'''

        # Selecting Items From Database For Treeview
        self.cursor.execute("SELECT  * FROM ITEMS")

        # Rows Of Data
        self.rows = self.cursor.fetchall()

        # Deleting Table ITEMS
        # self.conn.execute("DROP TABLE ITEMS")

        # Executing SQLITE Commands
        self.cursor.execute(sql)

        # Inserting Data Into Treeview
        self.count = 0
        self.nu = 1

        # Through For Loop
        for a, b, c, d, e, f, g in self.rows:
            # IF ELSE Statements For Striped Treeview
            if self.count % 2 == 0:
                self.item_treeview.insert('', END,
                                          values=(self.nu, e, f, g, d, c, b),
                                          tags=('evenrow',))
                self.count += 1
                self.nu += 1
            else:
                self.item_treeview.insert('', END,
                                          values=(self.nu, e, f, g, d, c, b),
                                          tags=('oddrow',))
                self.count += 1
                self.nu += 1

        # Packing Vertical Scrollbar
        self.vsb.pack(side=RIGHT, fill=Y)

        # Packing Horizontal Scrollbar
        self.hsb.pack(side=BOTTOM, fill=X)

        # Configuring Scrollbars For Treeview
        self.vsb.configure(command=self.item_treeview.yview)
        self.hsb.configure(command=self.item_treeview.xview)

        # Packing Item Treeview
        self.item_treeview.pack(fill=BOTH, expand=1)

        # Add Item Button
        self.add_item_btn = ttk.Button(self.item_menu, text="Add Item", compound=LEFT,
                                       cursor="hand2", style="S.TButton", command=self.add_item_to_menu)
        self.add_item_btn.pack(anchor=W)

    # Adding Items To Particulars Treeview from Item Treeview
    def add_item_to_menu(self):

        # Getting The Last Row Of Particulars Treeview
        last_row = len(self.particulars_treeview.get_children())

        print(last_row)

        # Giving Variable For Focusing Treeview
        cursor_row = self.item_treeview.focus()

        # Using Focus Variable To Get Item Values
        contents = self.item_treeview.item(cursor_row)

        # Variable For Storing Item Values
        row = contents['values']

        # Variables For Getting Row Values
        name = row[1]
        rate = row[2]
        qty = row[3]
        item_code = row[4]
        sub_category = row[5]
        category = row[6]

        # IF Statement For Checking If Name Already Exist in Order
        if name in self.items[category].keys():
            tmsg.showinfo("Error", "Item already exist in your order")
            return

        # IF Statement For Checking If Quantity Is Digit in Order
        '''if not qty.isdigit():
            tmsg.showinfo("Error", "Please Enter Valid Quantity")
            return'''

        # Getting Items in List Variable
        self.lis = [name, rate, qty, str(int(rate) * int(qty)), item_code, sub_category, category]
        self.items[category][name] = self.lis
        # IF ELSE Statement For Getting Last Row And Inserting It To Particulars Treeview
        if last_row != 0:
            print(str(1))
            if last_row % 2 != 0:
                self.particulars_treeview.insert('', END,
                                                 values=(
                                                     str(last_row + 1), name, rate, qty,
                                                     item_code, sub_category, category),
                                                 tags=('oddrow',))
            else:
                self.particulars_treeview.insert('', END,
                                                 values=(
                                                     str(last_row + 1), name, rate, qty,
                                                     item_code, sub_category, category),
                                                 tags=('evenrow',))
            self.update_total_price()
        else:
            self.particulars_treeview.insert('', END,
                                             values=(1, name, rate, qty, item_code, sub_category, category),
                                             tags=('oddrow',))
            self.update_total_price()


# Testing Invoice Class
# Root Window
root = Tk()


def get_location(event):
    x = root.winfo_rootx()
    y = root.winfo_rooty()
    return print("X:- " + str(x) + " Y:- " + str(y))


# Invoice Class
obj = Invoice(root)
# Loop For Running Root Window
root.mainloop()
