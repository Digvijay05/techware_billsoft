import json
import re
import sqlite3
from datetime import date, timedelta
from threading import Thread
import time
import openpyxl.drawing.image
from openpyxl import load_workbook, Workbook

self = Thread()
self.start()
# Connecting To Database
self.conn = sqlite3.connect("DB\\Business.db")
self.cursor = self.conn.cursor()

# Executing Commands In Cursor
self.cursor.execute("SELECT * FROM CATEGORY")

# Rows of Data
self.rows = self.cursor.fetchall()

# Initializing Items List
self.items = {}

# Through For Loop
for id, category in self.rows:
    # Appending To Items List
    self.items[category] = {}

# Initializing Categories List
self.categories = {}

# Through For Loop
for id, category in self.rows:
    # Appending To Categories List
    self.categories[category] = {}

self.conn = sqlite3.connect("DB\\Invoices.db")
self.cursor = self.conn.cursor()
self.cursor.execute("SELECT * FROM INVOICES WHERE Invoice_No = ?", ("WF1001",))
self.rows = self.cursor.fetchall()
self.t = time.localtime(time.time())
self.reports_workbook = Workbook()
self.reports_sheet = self.reports_workbook.active
self.reports_sheet.title = f"{self.t.tm_mon}-{self.t.tm_year} Reports"
self.row = 0
self.col = 0
if self.rows == [] or self.rows == None:
    print("no data")
else:
    print(self.rows)
    self.cols = ["No.", "Name", "Phone Number"]
    for i in self.categories:
        self.cols.append(i)
    # Giving Column names in Reports Sheet
    # Appending Column Name To No.
    self.reports_sheet.cell(row=self.row + 1, column=self.col + 1,
                            value=self.cols[0])
    # Appending Column Name To Name
    self.reports_sheet.cell(row=self.row + 1, column=self.col + 1, value=self.cols[1])
    # Appending Column Name To Phone No.
    self.reports_sheet.cell(row=self.row + 1, column=self.col + 1, value=self.rows[6])
    # row_cat = category to append in row
    for row_cat in self.categories:
        # Adding Categories in Reports Dynamically
        self.reports_sheet.cell(row=self.row + 1, column=self.col + self.reports_sheet.max_column,
                                value=row_cat)

    self.reports_workbook.save(f"{self.t.tm_mon},{self.t.tm_year}.xlsx")
    self.reports_workbook.close()
